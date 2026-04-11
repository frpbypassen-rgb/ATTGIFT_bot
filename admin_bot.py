import telebot
from telebot import types
from pymongo import MongoClient
import datetime
import certifi
import random
import string
import os
from flask import Flask
import threading
import time
import io
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# إعدادات بوت الإدارة
# ==========================================
ADMIN_BOT_TOKEN = "8628847468:AAE2-SDgYkEzQbGluTEAVF_rmzxwdlNVfIw"
CUSTOMER_BOT_TOKEN = "8769145956:AAEKIAKJ2sGn9HFu_-M8diyND1J754fp_Wc" # لكي يرسل إشعارات للعملاء

OWNER_ID = 1262656649
ADMIN_IDS = [OWNER_ID] 

MONGO_URI = "mongodb+srv://frpbypassen_db_user:LpovkVYkrNU7qePp@attgift.rdamxpj.mongodb.net/?retryWrites=true&w=majority&appName=ATTGIFT"
client = MongoClient(MONGO_URI, tlsCAFile=certifi.where(), maxPoolSize=50, connectTimeoutMS=5000)
db = client["AlAhram_DB"]

users = db["users"]
stock = db["stock"]
cards = db["cards"]
transactions = db["transactions"]
counters = db["counters"]
admins_db = db["admins"]

# تشغيل البوت الأساسي للإدارة، وتهيئة بوت العملاء كـ "ساعي بريد"
bot = telebot.TeleBot(ADMIN_BOT_TOKEN)
client_notifier = telebot.TeleBot(CUSTOMER_BOT_TOKEN)

MENU_BUTTONS = [
    "👥 المستخدمين", "⚙️ إدارة عميل", "🎫 توليد", "➕ منتج", 
    "💳 شحن يدوي", "💰 ضبط الرصيد", "🧾 سجل الفواتير", 
    "📦 إدارة المخزون", "📊 تقارير إكسيل", "💵 أسعار المستويات"
]

temp_admin_data = {}

# ==========================================
# الدوال المساعدة والإكسيل
# ==========================================
def safe_str(text):
    if text is None: return "بدون"
    res = str(text)
    for char in ["_", "*", "`", "[", "]", "(", ")", "~", ">", "#", "+", "-", "=", "|", "{", "}", ".", "!"]:
        res = res.replace(char, " ")
    return res.strip()

def is_admin(uid):
    if uid in ADMIN_IDS: return True
    if admins_db.find_one({"_id": uid}): return True
    return False

def find_customer(text):
    text = text.strip()
    if text.isdigit():
        u = users.find_one({"_id": int(text)})
        if u: return u
    clean_phone = text.replace("+", "").replace(" ", "").lstrip("0")
    if clean_phone:
        u = users.find_one({"phone": {"$regex": f"{clean_phone}$"}})
        if u: return u
    return None

def generate_admin_report_excel(report_type, history_data, summary_data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    if report_type == "all":
        ws.title = "سجل المبيعات الشامل"
        ws.append(["رقم الفاتورة", "التاريخ", "الاسم", "الهاتف", "نوع العملية", "البيان", "الكمية", "المبلغ"])
        for h in history_data:
            ws.append([h.get("order_id", "-"), h.get("date", ""), h.get("user_name", ""), h.get("phone", ""), h.get("type", ""), h.get("item_name", ""), h.get("quantity", 1), h.get("price", h.get("amount", 0))])
        if summary_data:
            ws_sum = wb.create_sheet(title="ملخص العملاء")
            ws_sum.append(["اسم العميل", "رقم الهاتف", "إجمالي المشتريات"])
            for phone, data in summary_data.items():
                ws_sum.append([data["name"], phone, data["spent"]])

    elif report_type == "single":
        ws.title = "كشف حساب عميل"
        ws.append(["رقم الفاتورة", "التاريخ", "نوع العملية", "البيان", "الكمية", "المبلغ"])
        total_in, total_out = 0, 0
        for h in history_data:
            price_or_amount = h.get("price", h.get("amount", 0))
            ws.append([h.get("order_id", "-"), h.get("date", ""), h.get("type", ""), h.get("item_name", "-"), h.get("quantity", "-"), price_or_amount])
            if h.get("type") == "شراء": total_out += price_or_amount
            else: total_in += price_or_amount
        ws.append([]); ws.append(["", "", "", "إجمالي الإيداعات:", "", total_in])
        ws.append(["", "", "", "إجمالي المشتريات:", "", total_out])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = white_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            max_len = 0; column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column].width = max_len + 4

    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return stream

def generate_products_template():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["القسم", "الفئة", "الاسم", "السعر 1", "السعر 2", "السعر 3", "كود الشحن", "الرقم التسلسلي", "الرقم السري (PIN)", "اوبريشن كود"])
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return stream

def generate_simple_excel(data_list, title):
    wb = openpyxl.Workbook(); ws = wb.active
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"); header_font = Font(color="FFFFFF", bold=True)
    ws.append(["القسم", "الفئة", "الاسم", "السعر 1", "السعر 2", "السعر 3", "كود الشحن", "الرقم التسلسلي", "الرقم السري (PIN)", "اوبريشن كود"])
    for col_num in range(1, 11):
        cell = ws.cell(row=1, column=col_num); cell.font = header_font; cell.fill = header_fill
    for d in data_list:
        ws.append([d.get('category', ''), d.get('subcategory', ''), d.get('name', ''), d.get('price_1', 0), d.get('price_2', 0), d.get('price_3', 0), str(d.get('code', '')), str(d.get('serial', '')), str(d.get('pin', '')), str(d.get('op_code', ''))])
    for col in ws.columns:
        max_len = 0; column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_len + 4
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return stream

def process_radar_logic(chat_id, raw_codes_list, product_info=None):
    if not raw_codes_list: return bot.send_message(chat_id, "❌ لم يتم العثور على بيانات صالحة.")
    bot.send_message(chat_id, "⏳ يتم فحص التكرار في قاعدة البيانات...")
    
    incoming = [str(item['code']).strip() for item in raw_codes_list if 'code' in item]
    db_set = set([str(doc['code']).strip() for doc in stock.find({"code": {"$in": incoming}})])
        
    acc_docs, rej_docs, seen = [], [], set()
    for item in raw_codes_list:
        c = str(item.get('code', '')).strip()
        if not c: continue
        if c in db_set or c in seen: rej_docs.append(item)
        else:
            seen.add(c)
            new_doc = item.copy()
            if product_info: new_doc.update(product_info)
            new_doc["sold"] = False; new_doc["added_at"] = datetime.datetime.now()
            acc_docs.append(new_doc)
            
    if acc_docs: stock.insert_many(acc_docs)
    bot.send_message(chat_id, f"📊 **النتيجة:**\n✅ مقبولة: `{len(acc_docs)}`\n❌ مكررة ومرفوضة: `{len(rej_docs)}`", parse_mode="Markdown")
    
    try:
        if acc_docs:
            f_acc = generate_simple_excel(acc_docs, "Accepted")
            bot.send_document(chat_id, ("Accepted.xlsx", f_acc.getvalue()), caption="📁 الأكواد المقبولة.")
        if rej_docs:
            f_rej = generate_simple_excel(rej_docs, "Rejected")
            bot.send_document(chat_id, ("Rejected.xlsx", f_rej.getvalue()), caption="⚠️ الأكواد المكررة (مرفوضة).")
    except Exception as e:
        bot.send_message(chat_id, f"⚠️ تم التحديث، لكن فشل إرسال الملف: {e}")

def admin_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("👥 المستخدمين", "⚙️ إدارة عميل")
    kb.add("🎫 توليد", "💳 شحن يدوي")
    kb.add("💰 ضبط الرصيد", "➕ منتج")
    kb.add("📦 إدارة المخزون", "🧾 سجل الفواتير") 
    kb.add("📊 تقارير إكسيل", "💵 أسعار المستويات") 
    return kb

# ==========================================
# الأوامر الأساسية ولوحة التحكم
# ==========================================
@bot.message_handler(commands=['start', 'admin'])
def show_admin_dashboard(msg):
    if not is_admin(msg.chat.id):
        bot.send_message(msg.chat.id, "❌ ليس لديك صلاحية للوصول للوحة التحكم.")
        return
    bot.send_message(msg.chat.id, "👑 مرحباً بك في لوحة تحكم شركة الأهرام.", reply_markup=admin_menu())

@bot.message_handler(func=lambda m: m.text == "👥 المستخدمين")
def admin_users_list(msg):
    if not is_admin(msg.chat.id): return
    try:
        text = "👥 قائمة آخر 30 مستخدم مسجل:\n\n"
        for u in users.find().sort("join", -1).limit(30):
            stat = "✅" if u.get('status') == 'active' else ("🚫" if u.get('status') == 'blocked' else "❄️")
            text += f"📛 {safe_str(u.get('name'))} | 📱 {u.get('phone', 'بدون')} | 💰 {round(float(u.get('balance', 0)), 2)} | {stat}\n"
        bot.send_message(msg.chat.id, text)
    except Exception as e: bot.send_message(msg.chat.id, f"❌ خطأ: {e}")

@bot.message_handler(func=lambda m: m.text == "🧾 سجل الفواتير")
def admin_invoices_log(msg):
    if not is_admin(msg.chat.id): return
    history = list(transactions.find({"type": "شراء", "order_id": {"$exists": True}}).sort("_id", -1).limit(40))
    if not history: return bot.send_message(msg.chat.id, "لا توجد فواتير.")
    text = "🧾 سجل آخر 40 فاتورة مبيعات:\n\n"
    for t in history:
        text += f"▪️ #{t.get('order_id', '-')} | 👤 {safe_str(t.get('user_name'))} | 🛒 {t.get('item_name', '-')} (x{t.get('quantity', 1)}) | 💰 {round(float(t.get('price', 0)), 2)}\n"
    bot.send_message(msg.chat.id, text)

# ==========================================
# إدارة العملاء
# ==========================================
@bot.message_handler(func=lambda m: m.text == "⚙️ إدارة عميل")
def admin_manage_customer(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "يرجى إرسال رقم هاتف العميل:")
    bot.register_next_step_handler(ask, admin_find_and_render_customer)

def admin_find_and_render_customer(msg):
    if msg.text in MENU_BUTTONS: return
    u = find_customer(msg.text)
    if not u: return bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    admin_render_customer_card(u["_id"], msg.chat.id)

def admin_render_customer_card(uid, chat_id, message_id=None):
    u = users.find_one({"_id": uid})
    if not u: return
    
    t_lbl = "مستوى 1 🥉" if u.get("tier", 1) == 1 else ("مستوى 2 🥈" if u.get("tier", 1) == 2 else "مستوى 3 🥇")
    s_lbl = "محظور 🚫" if u.get("status") == "blocked" else ("نشط ✅" if u.get("status") == "active" else "مجمد ❄️")
    
    info = f"👤 **ملف بيانات العميل:**\n\n📛 الاسم: {safe_str(u.get('name'))}\n🆔 ID: `{uid}`\n📱 الهاتف: `{u.get('phone')}`\n💰 الرصيد: **{round(float(u.get('balance',0)),2)}**\n🎚️ المستوى: **{t_lbl}**\nحالة الحساب: **{s_label}**"
    
    kb = types.InlineKeyboardMarkup(row_width=2)
    if u.get("status") == "active": kb.add(types.InlineKeyboardButton("❄️ تجميد الحساب", callback_data=f"adm_freeze_{uid}"), types.InlineKeyboardButton("🚫 حظر", callback_data=f"adm_block_{uid}"))
    elif u.get("status") == "frozen": kb.add(types.InlineKeyboardButton("✅ تفعيل الحساب", callback_data=f"adm_activate_{uid}"), types.InlineKeyboardButton("🚫 حظر", callback_data=f"adm_block_{uid}"))
    elif u.get("status") == "blocked": kb.add(types.InlineKeyboardButton("✅ فك الحظر", callback_data=f"adm_activate_{uid}"))
    kb.add(types.InlineKeyboardButton("🎚️ تغيير المستوى", callback_data=f"adm_chgtier_{uid}"), types.InlineKeyboardButton("📊 تقرير العمليات", callback_data=f"adm_report_{uid}"))
    
    if message_id:
        try: bot.edit_message_text(info, chat_id, message_id, reply_markup=kb, parse_mode="Markdown")
        except: pass
    else: bot.send_message(chat_id, info, reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("adm_"))
def handle_admin_customer_actions(call):
    action, uid = call.data.split("_")[1], int(call.data.split("_")[2])
    if action == "report":
        h = list(transactions.find({"uid": uid}).sort("_id", -1).limit(15))
        if not h: return bot.answer_callback_query(call.id, "لا توجد عمليات.")
        txt = f"📊 آخر عمليات العميل:\n\n"
        for t in h: txt += f"▪️ {t.get('date', '')} | {t.get('type', '')} | {round(float(t.get('price', t.get('amount', 0))), 2)}\n"
        bot.send_message(call.message.chat.id, txt); bot.answer_callback_query(call.id); return

    if action == "chgtier":
        kb = types.InlineKeyboardMarkup(row_width=3)
        kb.add(types.InlineKeyboardButton("المستوى 1 🥉", callback_data=f"set_tier_1_{uid}"), types.InlineKeyboardButton("المستوى 2 🥈", callback_data=f"set_tier_2_{uid}"), types.InlineKeyboardButton("المستوى 3 🥇", callback_data=f"set_tier_3_{uid}"))
        kb.add(types.InlineKeyboardButton("🔙 رجوع", callback_data=f"adm_back_{uid}"))
        return bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=kb)

    if action == "freeze": users.update_one({"_id": uid}, {"$set": {"status": "frozen"}})
    elif action == "activate":
        users.update_one({"_id": uid}, {"$set": {"status": "active", "failed_attempts": 0}})
        try: client_notifier.send_message(uid, "✅ تم تفعيل حسابك بنجاح من قبل الإدارة. يمكنك الاستمتاع بخدمات المتجر الآن.")
        except: pass
    elif action == "block":
        users.update_one({"_id": uid}, {"$set": {"status": "blocked"}})
        try: client_notifier.send_message(uid, "تم حظرك من قبل الادارة وللاستفسار تواصل معنا")
        except: pass
    
    bot.answer_callback_query(call.id, "تم تحديث بيانات العميل بنجاح.")
    admin_render_customer_card(uid, call.message.chat.id, call.message.message_id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("set_tier_"))
def handle_tier_setting(call):
    lvl, uid = int(call.data.split("_")[2]), int(call.data.split("_")[3])
    users.update_one({"_id": uid}, {"$set": {"tier": lvl}})
    bot.answer_callback_query(call.id, f"تم تغيير المستوى إلى {lvl}")
    admin_render_customer_card(uid, call.message.chat.id, call.message.message_id)

# ==========================================
# إدارة المخزون والمنتجات (الإضافة والتعديل)
# ==========================================
@bot.message_handler(func=lambda m: m.text == "➕ منتج")
def admin_add_product(msg):
    if not is_admin(msg.chat.id): return
    txt = "➕ **إضافة منتجات للمخزن:**\n\n📁 **الطريقة الأولى:** رفع ملف الإكسيل المرفق.\n✍️ **الطريقة الثانية:** إرسال `القسم:الفئة:الاسم:السعر1:السعر2:السعر3`"
    try:
        temp_file = generate_products_template()
        ask = bot.send_document(msg.chat.id, ("Template.xlsx", temp_file.getvalue()), caption=txt, parse_mode="Markdown")
        bot.register_next_step_handler(ask, admin_process_product_input)
    except Exception as e: bot.send_message(msg.chat.id, f"❌ خطأ: {e}")

def admin_process_product_input(msg):
    if msg.text in MENU_BUTTONS: return
    if msg.document:
        if not msg.document.file_name.endswith('.xlsx'): return bot.send_message(msg.chat.id, "❌ يرجى رفع ملف .xlsx فقط.")
        bot.send_message(msg.chat.id, "⏳ جاري القراءة...")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(bot.download_file(bot.get_file(msg.document.file_id).file_path)))
            ws = wb.active
            batch, errs = [], 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or not row[0]: continue
                try:
                    p1 = float(row[3])
                    batch.append({
                        "category": str(row[0]).strip(), "subcategory": str(row[1]).strip(), "name": str(row[2]).strip(),
                        "price_1": p1, "price_2": float(row[4]) if len(row)>4 and row[4] else p1, "price_3": float(row[5]) if len(row)>5 and row[5] else p1,
                        "code": str(row[6]).strip() if len(row)>6 and row[6] else "", "serial": str(row[7]).strip() if len(row)>7 and row[7] else "",
                        "pin": str(row[8]).strip() if len(row)>8 and row[8] else "", "op_code": str(row[9]).strip() if len(row)>9 and row[9] else ""
                    })
                except: errs += 1
            if errs > 0: bot.send_message(msg.chat.id, f"⚠️ تم تخطي {errs} صف لخطأ بالبيانات.")
            process_radar_logic(msg.chat.id, batch)
        except Exception as e: bot.send_message(msg.chat.id, f"❌ خطأ: {e}")
    elif msg.text and ":" in msg.text:
        try:
            p = msg.text.split(":")
            p1 = float(p[3].strip())
            temp_admin_data[msg.chat.id] = {"info": {"cat": p[0].strip(), "sub": p[1].strip(), "name": p[2].strip(), "price_1": p1, "price_2": float(p[4].strip()) if len(p)>4 else p1, "price_3": float(p[5].strip()) if len(p)>5 else p1}}
            ask = bot.send_message(msg.chat.id, "✅ تم الحفظ. أرسل الأكواد الآن:\n`الكود:التسلسلي:PIN:أوبريشن`", parse_mode="Markdown")
            bot.register_next_step_handler(ask, admin_process_manual_codes)
        except: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

def admin_process_manual_codes(msg):
    if msg.text in MENU_BUTTONS: return
    p_info = temp_admin_data.get(msg.chat.id, {}).get("info")
    if not p_info: return
    batch = []
    if msg.text:
        for line in msg.text.split("\n"):
            if not line.strip(): continue
            pts = line.split(":")
            batch.append({"code": pts[0].strip(), "serial": pts[1].strip() if len(pts)>1 else "", "pin": pts[2].strip() if len(pts)>2 else "", "op_code": pts[3].strip() if len(pts)>3 else ""})
    process_radar_logic(msg.chat.id, batch, p_info)

@bot.message_handler(func=lambda m: m.text == "💵 أسعار المستويات")
def admin_bulk_price_edit(msg):
    if not is_admin(msg.chat.id): return
    prods = list(stock.aggregate([{"$match": {"sold": False}}, {"$group": {"_id": "$name", "cat": {"$first": "$category"}, "sub": {"$first": "$subcategory"}, "p1": {"$first": "$price_1"}, "p2": {"$first": "$price_2"}, "p3": {"$first": "$price_3"}}}]))
    if not prods: return bot.send_message(msg.chat.id, "❌ لا توجد منتجات متاحة.")
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(["الاسم", "القسم", "الفئة", "سعر 1", "سعر 2", "سعر 3"])
    for p in prods: ws.append([p["_id"], p.get("cat", ""), p.get("sub", ""), p.get("p1", 0), p.get("p2", 0), p.get("p3", 0)])
    for col in ws.columns:
        mx = 0; col_l = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > mx: mx = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_l].width = mx + 4
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    ask = bot.send_document(msg.chat.id, ("Prices.xlsx", stream.getvalue()), caption="📁 أداة تعديل الأسعار الشاملة.\nعدل وأعد الرفع.")
    bot.register_next_step_handler(ask, admin_finalize_price_update)

def admin_finalize_price_update(msg):
    if msg.text in MENU_BUTTONS: return
    if not msg.document or not msg.document.file_name.endswith('.xlsx'): return bot.send_message(msg.chat.id, "❌ يرجى رفع ملف إكسيل.")
    bot.send_message(msg.chat.id, "⏳ جاري التحديث...")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(bot.download_file(bot.get_file(msg.document.file_id).file_path))); ws = wb.active; count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]: continue
            try:
                p1 = float(row[3]) if row[3] is not None else 0
                p2 = float(row[4]) if row[4] is not None else p1
                p3 = float(row[5]) if row[5] is not None else p1
                res = stock.update_many({"name": str(row[0]).strip(), "sold": False}, {"$set": {"price_1": p1, "price_2": p2, "price_3": p3}})
                if res.modified_count > 0: count += 1
            except: pass
        bot.send_message(msg.chat.id, f"✅ تم تحديث أسعار `{count}` نوع منتج.")
    except Exception as e: bot.send_message(msg.chat.id, f"❌ خطأ: {e}")

@bot.message_handler(func=lambda m: m.text == "📦 إدارة المخزون")
def manage_stock_cmd(msg):
    if not is_admin(msg.chat.id): return
    names = stock.distinct("name", {"sold": False})
    if not names: return bot.send_message(msg.chat.id, "❌ المخزن فارغ حالياً.")
    text = "📦 **المنتجات المتوفرة حالياً:**\n\n"
    for n in names: text += f"▪️ `{n}` (الكمية: {stock.count_documents({'name': n, 'sold': False})})\n"
    ask = bot.send_message(msg.chat.id, text + "\n👉 أرسل **اسم المنتج** (لنسخه):", parse_mode="Markdown")
    bot.register_next_step_handler(ask, show_stock_item_panel)

def show_stock_item_panel(msg):
    if msg.text in MENU_BUTTONS: return
    name = msg.text.strip(); item = stock.find_one({"name": name, "sold": False})
    if not item: return bot.send_message(msg.chat.id, "❌ غير موجود.")
    p1 = item.get("price_1", item.get("price", 0))
    text = f"📦 المنتج: `{name}`\n💰 الأسعار:\nم1: {p1} | م2: {item.get('price_2', p1)} | م3: {item.get('price_3', p1)}\n📊 الكمية: {stock.count_documents({'name': name, 'sold': False})}"
    temp_admin_data[msg.chat.id] = {"mng_item_name": name}
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("💰 تعديل السعر", callback_data="stk_price"), types.InlineKeyboardButton("➖ حذف كمية", callback_data="stk_delqty"), types.InlineKeyboardButton("❌ مسح بالكامل", callback_data="stk_delall"))
    bot.send_message(msg.chat.id, text, reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("stk_"))
def handle_stock_actions(call):
    action = call.data.split("_")[1]
    name = temp_admin_data.get(call.message.chat.id, {}).get("mng_item_name")
    if not name: return bot.answer_callback_query(call.id, "❌ انتهت الجلسة.", show_alert=True)
    if action == "price":
        ask = bot.send_message(call.message.chat.id, f"أرسل الأسعار للمنتج `{name}` (سعر1:سعر2:سعر3):", parse_mode="Markdown")
        bot.register_next_step_handler(ask, admin_update_stock_price, name)
    elif action == "delqty":
        ask = bot.send_message(call.message.chat.id, f"أرسل عدد الأكواد لمسحها من `{name}`:", parse_mode="Markdown")
        bot.register_next_step_handler(ask, admin_delete_stock_qty, name)
    elif action == "delall":
        res = stock.delete_many({"name": name, "sold": False})
        bot.edit_message_text(f"✅ تم حذف {res.deleted_count} كود من المنتج `{name}` بالكامل.", call.message.chat.id, call.message.message_id, parse_mode="Markdown")

def admin_update_stock_price(msg, name):
    if msg.text in MENU_BUTTONS: return
    try:
        p = msg.text.split(":"); p1 = float(p[0].strip())
        stock.update_many({"name": name, "sold": False}, {"$set": {"price_1": p1, "price_2": float(p[1].strip()) if len(p)>1 else p1, "price_3": float(p[2].strip()) if len(p)>2 else p1}})
        bot.send_message(msg.chat.id, f"✅ تم تحديث الأسعار للمنتج `{name}` بنجاح.", parse_mode="Markdown")
    except: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

def admin_delete_stock_qty(msg, name):
    if msg.text in MENU_BUTTONS: return
    try:
        qty = int(msg.text.strip()); docs = list(stock.find({"name": name, "sold": False}).limit(qty))
        if not docs: return bot.send_message(msg.chat.id, "❌ لا يوجد رصيد متاح للحذف.")
        stock.delete_many({"_id": {"$in": [d['_id'] for d in docs]}})
        bot.send_message(msg.chat.id, f"✅ تم حذف {len(docs)} كود بنجاح.")
    except: bot.send_message(msg.chat.id, "❌ يجب إرسال أرقام صحيحة.")

# ==========================================
# استخراج التقارير والماليات
# ==========================================
@bot.message_handler(func=lambda m: m.text == "📊 تقارير إكسيل")
def admin_excel_reports(msg):
    if not is_admin(msg.chat.id): return
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("📄 تقرير شامل", callback_data="rep_all"), types.InlineKeyboardButton("👤 تقرير عميل", callback_data="rep_single"))
    bot.send_message(msg.chat.id, "الرجاء اختيار نوع التقرير:", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("rep_"))
def admin_report_dates(call):
    temp_admin_data[call.message.chat.id] = {"rep_type": call.data.split("_")[1]}
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("🕒 كل الأوقات", callback_data="dt_all"), types.InlineKeyboardButton("📅 تحديد فترة", callback_data="dt_custom"))
    bot.edit_message_text("تحديد الفترة الزمنية:", call.message.chat.id, call.message.message_id, reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("dt_"))
def admin_report_finalize(call):
    uid = call.message.chat.id; mode = call.data.split("_")[1]; rep_type = temp_admin_data.get(uid, {}).get("rep_type")
    if not rep_type: return bot.answer_callback_query(call.id, "انتهت الجلسة.", show_alert=True)
    if mode == "all": admin_execute_report(uid, rep_type, None, None)
    else:
        ask = bot.send_message(uid, "أرسل التاريخ (من-إلى) YYYY-MM-DD:YYYY-MM-DD :", parse_mode="Markdown")
        bot.register_next_step_handler(ask, admin_process_date_input, rep_type)

def admin_process_date_input(msg, rep_type):
    if msg.text in MENU_BUTTONS: return
    try:
        p = msg.text.split(":")
        admin_execute_report(msg.chat.id, rep_type, p[0].strip(), p[1].strip())
    except: bot.send_message(msg.chat.id, "❌ تنسيق غير صحيح.")

def admin_execute_report(uid, rep_type, start_date, end_date):
    filtr = {}
    if start_date and end_date: filtr = {"date": {"$gte": start_date, "$lte": end_date + " 23:59"}}
    if rep_type == "all":
        bot.send_message(uid, "⏳ جاري تجميع بيانات المبيعات...")
        data = list(transactions.find(filtr).sort("_id", -1))
        if not data: return bot.send_message(uid, "❌ لا توجد عمليات.")
        summary = {}
        for t in data:
            if t.get("type") == "شراء":
                phone = t.get("phone", "بدون")
                if phone not in summary: summary[phone] = {"name": t.get("user_name", "غير مسجل"), "spent": 0}
                summary[phone]["spent"] += t.get("price", 0)
        file_stream = generate_admin_report_excel("all", data, summary)
        try: bot.send_document(uid, ("Report_All.xlsx", file_stream.getvalue()), caption="✅ التقرير الشامل.")
        except Exception as e: bot.send_message(uid, f"❌ خطأ الإرسال: {e}")
    else:
        ask = bot.send_message(uid, "يرجى إرسال رقم الهاتف للعميل المطلوب استخراج تقريره:")
        bot.register_next_step_handler(ask, admin_execute_single_report, filtr)

def admin_execute_single_report(msg, filtr):
    if msg.text in MENU_BUTTONS: return
    u = find_customer(msg.text)
    if not u: return bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    query = {"uid": u["_id"]}; query.update(filtr)
    data = list(transactions.find(query).sort("_id", -1))
    if not data: return bot.send_message(msg.chat.id, "❌ لا توجد عمليات لهذا العميل.")
    file_stream = generate_admin_report_excel("single", data)
    try: bot.send_document(msg.chat.id, (f"Report_{u.get('phone', 'Client')}.xlsx", file_stream.getvalue()), caption=f"✅ تقرير العميل: {safe_str(u.get('name'))}")
    except: bot.send_message(msg.chat.id, "❌ خطأ بالإرسال.")

@bot.message_handler(func=lambda m: m.text == "💰 ضبط الرصيد")
def admin_set_balance(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل التعديل (رقم_الهاتف:الرصيد_الجديد):\nمثال: `0940719000:500`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_set_balance)

def execute_set_balance(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        u = find_customer(parts[0])
        if u:
            new_bal = round(float(parts[1].strip()), 2)
            users.update_one({"_id": u["_id"]}, {"$set": {"balance": new_bal}})
            transactions.insert_one({"uid": u["_id"], "user_name": u.get("name"), "type": "ضبط رصيد من الإدارة", "amount": new_bal, "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
            bot.send_message(msg.chat.id, f"✅ تم ضبط رصيد العميل ليصبح {new_bal}")
            try: client_notifier.send_message(u["_id"], f"⚙️ إشعار من الإدارة: تم تحديث رصيدك ليصبح {new_bal}")
            except: pass
        else: bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    except: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

@bot.message_handler(func=lambda m: m.text == "💳 شحن يدوي")
def admin_direct_charge(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل الشحن (رقم_الهاتف:قيمة_الإضافة):\nمثال: `0940719000:50`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_direct_charge)

def execute_direct_charge(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        u = find_customer(parts[0])
        if u:
            amt = round(float(parts[1].strip()), 2)
            users.update_one({"_id": u["_id"]}, {"$inc": {"balance": amt}})
            transactions.insert_one({"uid": u["_id"], "user_name": u.get("name"), "type": "إضافة رصيد مباشر", "amount": amt, "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
            bot.send_message(msg.chat.id, f"✅ تمت الإضافة بنجاح.")
            try: client_notifier.send_message(u["_id"], f"🎁 إشعار من الإدارة: تم إيداع {amt} في حسابك.")
            except: pass
        else: bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    except: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

@bot.message_handler(func=lambda m: m.text == "🎫 توليد")
def admin_generate_cards(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل (العدد:القيمة)\nمثال لتوليد 10 كروت فئة 50:\n`10:50`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_generate_cards)

def execute_generate_cards(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        count = int(parts[0].strip())
        val = round(float(parts[1].strip()), 2)
        arr = []
        txt = f"✅ تم توليد {count} كروت شحن فئة {val}:\n\n"
        for _ in range(count):
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=12))
            arr.append({"code": code, "value": val, "used": False})
            txt += f"`{code}`\n"
        cards.insert_many(arr)
        bot.send_message(msg.chat.id, txt, parse_mode="Markdown")
    except: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

# ==========================================
# الأوامر المخفية
# ==========================================
@bot.message_handler(commands=['FRP', 'frp'])
def frp_reset_command(msg):
    if msg.chat.id != OWNER_ID: return
    ask = bot.send_message(msg.chat.id, "⚠️ **فورمات المصنع** ⚠️\n\nأرسل هذه الجملة حرفياً للتأكيد:\n`تأكيد الحذف النهائي`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_frp_wipe)

def execute_frp_wipe(msg):
    if msg.text == "تأكيد الحذف النهائي":
        users.delete_many({}); stock.delete_many({}); transactions.delete_many({}); admins_db.delete_many({}); counters.delete_many({}); cards.delete_many({})
        bot.send_message(msg.chat.id, "✅ تم تصفير النظام وعودته لحالة المصنع.")
    else: bot.send_message(msg.chat.id, "❌ تم إلغاء المسح.")

@bot.message_handler(commands=['ADD', 'add'])
def add_sub_admin_command(msg):
    if msg.chat.id != OWNER_ID: return
    ask = bot.send_message(msg.chat.id, "أرسل رقم الـ ID للترقية إلى مدير:")
    bot.register_next_step_handler(ask, execute_add_admin)

def execute_add_admin(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        new_admin_id = int(msg.text.strip())
        admins_db.update_one({"_id": new_admin_id}, {"$set": {"added_by": msg.chat.id, "join_date": datetime.datetime.now()}}, upsert=True)
        bot.send_message(msg.chat.id, f"✅ تمت ترقية `{new_admin_id}`.", parse_mode="Markdown")
        try: bot.send_message(new_admin_id, "🎉 تمت ترقيتك لمشرف في النظام. أرسل /admin للدخول.")
        except: pass
    except: bot.send_message(msg.chat.id, "❌ يرجى إرسال ID صحيح.")

@bot.message_handler(commands=['Block', 'block'])
def block_user_command(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل رقم الهاتف المراد حظره:")
    bot.register_next_step_handler(ask, execute_block)

def execute_block(msg):
    if msg.text in MENU_BUTTONS: return
    u = find_customer(msg.text)
    if u:
        users.update_one({"_id": u["_id"]}, {"$set": {"status": "blocked"}})
        bot.send_message(msg.chat.id, "✅ تم الحظر.")
        try: client_notifier.send_message(u["_id"], "تم حظرك من قبل الادارة وللاستفسار تواصل معنا", reply_markup=types.ReplyKeyboardRemove())
        except: pass
    else: bot.send_message(msg.chat.id, "❌ العميل غير موجود.")

# ==========================================
# تشغيل السيرفر 
# ==========================================
app = Flask(__name__)
@app.route('/')
def health_check(): return "<h1>Admin Bot is Running 🚀</h1>"

def bot_polling():
    bot.remove_webhook()
    while True:
        try: bot.infinity_polling(timeout=30, long_polling_timeout=15)
        except: time.sleep(5)

if __name__ == "__main__":
    threading.Thread(target=bot_polling, daemon=True).start()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
