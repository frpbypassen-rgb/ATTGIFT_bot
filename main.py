import telebot
from telebot import types
from pymongo import MongoClient, ReturnDocument
from bson import ObjectId
import datetime
import certifi
import random
import string
import os
from flask import Flask
import threading
import time
import requests 
import io
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# =======================================================
# 1. الإعدادات الأساسية (Configuration)
# =======================================================

API_TOKEN = "8769145956:AAEKIAKJ2sGn9HFu_-M8diyND1J754fp_Wc"

# المدير الأساسي (المالك)
OWNER_ID = 1262656649
ADMIN_IDS = [OWNER_ID] 

# الاتصال بقاعدة بيانات MongoDB مع تحسينات السرعة
MONGO_URI = "mongodb+srv://frpbypassen_db_user:LpovkVYkrNU7qePp@attgift.rdamxpj.mongodb.net/?retryWrites=true&w=majority&appName=ATTGIFT"
client = MongoClient(
    MONGO_URI, 
    tlsCAFile=certifi.where(),
    maxPoolSize=50, 
    connectTimeoutMS=5000
)
db = client["AlAhram_DB"]

# تعريف المجموعات (Collections)
users = db["users"]
stock = db["stock"]
cards = db["cards"]
transactions = db["transactions"]
counters = db["counters"]
admins_db = db["admins"]

# رابط جوجل شيت للمزامنة الخارجية (اختياري)
SHEET_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbzPrw8oANq8Aek6O6URoTU0kDVjb1ZtoVdYkhpqAqM6Nuws4ZmcPRC9JtoNZvWoMzUb/exec"

bot = telebot.TeleBot(API_TOKEN)

# قائمة الأزرار الرئيسية
MENU_BUTTONS = [
    "🛒 شراء", "💳 شحن", "👤 حسابي", "👥 المستخدمين", 
    "🎫 توليد", "➕ منتج", "💳 شحن يدوي", "⚙️ إدارة عميل", 
    "💰 ضبط الرصيد", "🧾 سجل الفواتير", "📦 إدارة المخزون", 
    "📊 تقارير إكسيل", "💵 أسعار المستويات", "🏪 العودة للمتجر"
]

temp_admin_data = {}

# =======================================================
# 2. الدوال المساعدة وتأمين النصوص (Utility Functions)
# =======================================================

def safe_str(text):
    """تنظيف النصوص من رموز الماركدوان لمنع تعليق الرسائل في تيليجرام"""
    if text is None:
        return "بدون"
    
    bad_chars = ["_", "*", "`", "[", "]", "(", ")", "~", ">", "#", "+", "-", "=", "|", "{", "}", ".", "!"]
    res = str(text)
    for char in bad_chars:
        res = res.replace(char, " ")
        
    return res.strip()

def is_admin(uid):
    """التحقق مما إذا كان المستخدم مديراً"""
    if uid in ADMIN_IDS:
        return True
        
    admin_record = admins_db.find_one({"_id": uid})
    if admin_record:
        return True
        
    return False

def get_all_admins():
    """جلب قائمة بكل المشرفين لإرسال الإشعارات"""
    admins = list(ADMIN_IDS)
    all_extra_admins = admins_db.find()
    
    for a in all_extra_admins:
        if a["_id"] not in admins:
            admins.append(a["_id"])
            
    return admins

def is_valid_val(val):
    """التحقق من صحة القيمة المدخلة في ملفات الإكسيل لمنع الأعمدة الفارغة"""
    if val is None:
        return False
        
    s_val = str(val).strip().lower()
    if s_val in ["", "none", "null", "بدون", "nan"]:
        return False
        
    return True

def find_customer(text):
    """البحث عن عميل عبر ID أو رقم الهاتف"""
    text = text.strip()
    
    # البحث بواسطة ID
    if text.isdigit():
        u = users.find_one({"_id": int(text)})
        if u:
            return u
            
    # البحث بواسطة رقم الهاتف
    clean_phone = text.replace("+", "").replace(" ", "").lstrip("0")
    if clean_phone:
        u = users.find_one({"phone": {"$regex": f"{clean_phone}$"}})
        if u:
            return u
            
    return None

def get_next_order_id():
    """توليد رقم تسلسلي للفواتير"""
    doc = counters.find_one_and_update(
        {"_id": "order_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    return doc["seq"]

def check_user_access(uid):
    """دالة شاملة للتحقق من صلاحية وصول العميل للنظام"""
    u = users.find_one({"_id": uid})
    
    if not u:
        bot.send_message(uid, "⚠️ النظام مغلق. أرسل /start للتسجيل أولاً.")
        return None
        
    if not u.get("name"):
        msg = bot.send_message(uid, "يرجى كتابة اسمك الكريم أولاً للاستمرار:")
        bot.register_next_step_handler(msg, process_name)
        return None
        
    if not u.get("phone"):
        bot.send_message(uid, "⚠️ يجب عليك مشاركة رقم هاتف حسابك أولاً.", reply_markup=contact_menu())
        return None
        
    if u.get("status") == "blocked":
        bot.send_message(uid, "تم حظرك من قبل الادارة وللاستفسار تواصل معنا")
        return None
        
    if u.get("status") != "active":
        bot.send_message(uid, "❌ حسابك قيد المراجعة. برجاء الانتظار حتى تفعيل الإدارة.")
        return None
        
    return u
    
# =======================================================
# 3. محرك معالجة الإكسيل والفواتير الديناميكية
# =======================================================

def generate_customer_excel_file(items_data, order_id, dt_now, product_name):
    """إنشاء فاتورة العميل الديناميكية التي تخفي الأعمدة الفارغة"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoice_{order_id}"
    
    # إعدادات الألوان والتنسيق
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    info_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    
    # فحص البيانات المتوفرة في هذه العملية لإظهار أو إخفاء الأعمدة
    has_serial = False
    has_pin = False
    has_opcode = False
    
    for d in items_data:
        if is_valid_val(d.get('serial')): has_serial = True
        if is_valid_val(d.get('pin')): has_pin = True
        if is_valid_val(d.get('op_code')): has_opcode = True
            
    # بناء الهيدر ديناميكياً
    headers = ["م", "اسم المنتج", "كود الشحن"]
    if has_serial: headers.append("الرقم التسلسلي")
    if has_pin: headers.append("الرقم السري (PIN)")
    if has_opcode: headers.append("أوبريشن كود")
        
    last_col_idx = len(headers)
    last_col_letter = get_column_letter(last_col_idx)
    
    # تصميم رأس الصفحة
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = "شركة الأهرام للإتصالات والتقنية"
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws['A1'].fill = title_fill
    ws['A1'].alignment = centered
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"فاتورة شراء رقم: #{order_id}   |   تاريخ العملية: {dt_now}"
    ws['A2'].fill = info_fill
    ws['A2'].alignment = centered
    ws.row_dimensions[2].height = 25
    
    ws.append([]) # سطر فارغ
    
    # إضافة عناوين الجدول
    ws.append(headers)
    for cell in ws[4]:
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = centered
        
    # تعبئة البيانات
    for i, d in enumerate(items_data, 1):
        row_data = [i, product_name, d['code']]
        if has_serial: row_data.append(d.get('serial', ''))
        if has_pin: row_data.append(d.get('pin', ''))
        if has_opcode: row_data.append(d.get('op_code', ''))
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.alignment = centered

    # ضبط تلقائي لعرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except Exception: pass
        ws.column_dimensions[column].width = max_length + 5

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # 👇 السطر الذي يحل المشكلة ويعرف الملف لدى تيليجرام 👇
    stream.name = f"Invoice_{order_id}.xlsx"
    return stream

def generate_admin_report_excel(report_type, history_data, summary_data=None):
    """إنشاء تقارير الإدارة الشاملة أو الفردية للعملاء"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    if report_type == "all":
        ws.title = "سجل المبيعات الشامل"
        headers = ["رقم الفاتورة", "التاريخ", "الاسم", "الهاتف", "نوع العملية", "البيان", "الكمية", "المبلغ"]
        ws.append(headers)
        
        for h in history_data:
            ws.append([
                h.get("order_id", "-"), h.get("date", ""), h.get("user_name", ""),
                h.get("phone", ""), h.get("type", ""), h.get("item_name", ""),
                h.get("quantity", 1), h.get("price", h.get("amount", 0))
            ])
            
        if summary_data:
            ws_sum = wb.create_sheet(title="ملخص العملاء")
            ws_sum.append(["اسم العميل", "رقم الهاتف", "إجمالي المشتريات"])
            for phone, data in summary_data.items():
                ws_sum.append([data["name"], phone, data["spent"]])

    elif report_type == "single":
        ws.title = "كشف حساب عميل"
        headers = ["رقم الفاتورة", "التاريخ", "نوع العملية", "البيان", "الكمية", "المبلغ"]
        ws.append(headers)
        
        total_in = 0
        total_out = 0
        for h in history_data:
            price_or_amount = h.get("price", h.get("amount", 0))
            ws.append([
                h.get("order_id", "-"), h.get("date", ""), h.get("type", ""),
                h.get("item_name", "-"), h.get("quantity", "-"), price_or_amount
            ])
            if h.get("type") == "شراء": total_out += price_or_amount
            else: total_in += price_or_amount
        
        ws.append([])
        ws.append(["", "", "", "إجمالي الإيداعات:", "", total_in])
        ws.append(["", "", "", "إجمالي المشتريات:", "", total_out])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # 👇 السطر الذي يحل المشكلة ويعرف الملف لدى تيليجرام 👇
    stream.name = f"Report_{int(time.time())}.xlsx"
    return stream


# =======================================================
# 4. رادار منع التكرار (Anti-Duplicate Radar)
# =======================================================

def process_radar_logic(chat_id, raw_codes_list, product_info=None):
    """فحص الأكواد المرفوعة وتوليد تقارير الإكسيل الفورية للقبول والرفض"""
    if not raw_codes_list:
        bot.send_message(chat_id, "❌ لم يتم العثور على أي بيانات صالحة لمعالجتها.")
        return
        
    bot.send_message(chat_id, "⏳ رادار الأمان يعمل الآن... يتم فحص التكرار في قاعدة البيانات.")
    
    # تجميع الأكواد المدخلة للبحث المجمع في قاعدة البيانات
    incoming_codes = []
    for item in raw_codes_list:
        if 'code' in item:
            incoming_codes.append(str(item['code']).strip())
            
    # البحث عن الأكواد التي لها وجود مسبق
    existing_in_db = stock.find({"code": {"$in": incoming_codes}})
    
    db_set = set()
    for doc in existing_in_db:
        db_set.add(str(doc['code']).strip())
        
    accepted_docs = []
    rejected_docs = []
    seen_in_batch = set()
    
    for item in raw_codes_list:
        current_code = str(item.get('code', '')).strip()
        if not current_code:
            continue
            
        # إذا كان الكود مكرراً في قاعدة البيانات أو مكرراً داخل نفس الملف المرفوع
        if current_code in db_set or current_code in seen_in_batch:
            rejected_docs.append(item)
        else:
            seen_in_batch.add(current_code)
            
            new_doc = item.copy()
            if product_info:
                new_doc.update(product_info)
                
            new_doc["sold"] = False
            new_doc["added_at"] = datetime.datetime.now()
            accepted_docs.append(new_doc)
            
    # تنفيذ الإضافة إلى قاعدة البيانات
    if len(accepted_docs) > 0:
        stock.insert_many(accepted_docs)
        
    # تقرير النتائج
    report_msg = (
        f"📊 **تقرير معالجة المخزون:**\n\n"
        f"✅ أكواد جديدة مقبولة: `{len(accepted_docs)}`\n"
        f"❌ أكواد مكررة مرفوضة: `{len(rejected_docs)}`"
    )
    bot.send_message(chat_id, report_msg, parse_mode="Markdown")
    
    # إرسال ملفات الإكسيل للفرز
    if len(accepted_docs) > 0:
        file_acc = generate_simple_excel(accepted_docs, "Accepted_New_Codes")
        bot.send_document(chat_id, file_acc, caption="📁 ملف الأكواد التي تم إضافتها بنجاح.")
        
    if len(rejected_docs) > 0:
        file_rej = generate_simple_excel(rejected_docs, "Rejected_Duplicates")
        bot.send_document(chat_id, file_rej, caption="⚠️ ملف الأكواد المكررة (التي تم استبعادها ومسحها).")

# =======================================================
# 5. القوائم (Menus)
# =======================================================

def menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🛒 شراء", "💳 شحن")
    kb.add("👤 حسابي")
    return kb

def contact_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(types.KeyboardButton("إرسال رقم الهاتف 📱", request_contact=True))
    return kb

def admin_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("👥 المستخدمين", "⚙️ إدارة عميل")
    kb.add("🎫 توليد", "💳 شحن يدوي")
    kb.add("💰 ضبط الرصيد", "➕ منتج")
    kb.add("📦 إدارة المخزون", "🧾 سجل الفواتير") 
    kb.add("📊 تقارير إكسيل", "💵 أسعار المستويات") 
    kb.add("🏪 العودة للمتجر")
    return kb

# =======================================================
# 6. التسجيل وإدارة العملاء (Registration & Onboarding)
# =======================================================

@bot.message_handler(commands=['start'])
def handle_start(msg):
    uid = msg.chat.id
    u = users.find_one({"_id": uid})

    if not u:
        new_user = {
            "_id": uid, 
            "name": None, 
            "phone": None, 
            "balance": 0.0, 
            "status": "frozen", 
            "tier": 1, 
            "failed_attempts": 0, 
            "join": datetime.datetime.now()
        }
        users.insert_one(new_user)
        u = users.find_one({"_id": uid})

    if not u.get("name"):
        ask_msg = bot.send_message(uid, "👋 مرحباً بك في شركة الأهرام للإتصالات والتقنية.\n\nيرجى البدء بكتابة **اسمك بالكامل** (أو اسم متجرك):")
        bot.register_next_step_handler(ask_msg, process_name)
    elif not u.get("phone"):
        safe_n = safe_str(u.get('name'))
        bot.send_message(uid, f"أهلاً بك يا {safe_n}! يرجى تزويدنا برقم هاتفك لاستكمال تفعيل الحساب.", reply_markup=contact_menu())
    else:
        status = u.get("status")
        if status == "blocked":
            bot.send_message(uid, "تم حظرك من قبل الادارة وللاستفسار تواصل معنا")
        elif status == "frozen":
            bot.send_message(uid, "حسابك الآن (قيد المراجعة). سيتم إشعارك فور تفعيل الحساب من قبل الإدارة.")
        else:
            bot.send_message(uid, "أهلاً بك مجدداً في المتجر 🏪", reply_markup=menu())

def process_name(msg):
    if not msg.text or msg.text.startswith('/'):
        ask_msg = bot.send_message(msg.chat.id, "يرجى كتابة اسم صحيح بدون رموز:")
        bot.register_next_step_handler(ask_msg, process_name)
        return
        
    name = msg.text.strip()
    users.update_one({"_id": msg.chat.id}, {"$set": {"name": name}})
    
    safe_n = safe_str(name)
    bot.send_message(
        msg.chat.id, 
        f"تشرفنا بك يا {safe_n}!\n\nالآن، يرجى مشاركة رقم هاتفك عبر الضغط على الزر بالأسفل 📱", 
        reply_markup=contact_menu()
    )

@bot.message_handler(content_types=['contact'])
def handle_contact_sharing(msg):
    uid = msg.chat.id
    
    if msg.contact.user_id != uid:
        bot.send_message(uid, "❌ يرجى مشاركة رقم هاتفك الخاص المرتبط بهذا الحساب.", reply_markup=contact_menu())
        return
        
    phone = msg.contact.phone_number
    if not phone.startswith('+'):
        phone = '+' + phone

    users.update_one({"_id": uid}, {"$set": {"phone": phone, "tier": 1}})
    
    u = users.find_one({"_id": uid})
    safe_n = safe_str(u.get("name"))
    
    if u.get("status") == "frozen":
        bot.send_message(
            uid, 
            "✅ تم استلام بياناتك بنجاح.\nحسابك الآن قيد المراجعة، يرجى الانتظار لحين التفعيل من قبل الإدارة.", 
            reply_markup=types.ReplyKeyboardRemove()
        )
        
        # إشعار الإدارة بالتسجيل الجديد
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("✅ تفعيل الحساب", callback_data=f"adm_activate_{uid}"),
               types.InlineKeyboardButton("🚫 حظر العميل", callback_data=f"adm_block_{uid}"))
               
        for admin_id in get_all_admins():
            try:
                bot.send_message(
                    admin_id, 
                    f"🆕 **طلب تسجيل جديد!**\n\n📛 الاسم: {safe_n}\n🆔 ID: `{uid}`\n📱 الهاتف: `{phone}`", 
                    reply_markup=kb
                )
            except Exception:
                pass
    else:
        bot.send_message(uid, "✅ حسابك نشط وجاهز.", reply_markup=menu())

# =======================================================
# 7. واجهة العميل الخاصة (Account Details)
# =======================================================

@bot.message_handler(func=lambda m: m.text == "👤 حسابي")
def show_account_info(msg):
    uid = msg.chat.id
    u = check_user_access(uid)
    if not u:
        return

    safe_n = safe_str(u.get('name'))
    bal = u.get('balance', 0.0)
    phone = u.get('phone', 'غير متوفر')
    status = "نشط ✅" if u.get('status') == 'active' else "قيد المراجعة ❄️"
    
    # لاحظ أننا لا نظهر المستوى للعميل بناءً على طلبك
    text = (
        f"👤 **بيانات حسابك**\n\n"
        f"📛 الاسم: {safe_n}\n"
        f"🆔 معرفك: `{uid}`\n"
        f"📱 الهاتف: `{phone}`\n"
        f"💰 رصيدك الحالي: **{bal}**\n"
        f"حالة الحساب: {status}"
    )
    
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(
        types.InlineKeyboardButton("🛒 سجل المشتريات", callback_data="client_purchases"),
        types.InlineKeyboardButton("🧾 كشف حساب تفصيلي", callback_data="client_statement")
    )
    bot.send_message(uid, text, parse_mode="Markdown", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data in ["client_purchases", "client_statement"])
def handle_client_reports(call):
    uid = call.message.chat.id
    u = check_user_access(uid)
    if not u:
        bot.answer_callback_query(call.id, "عذراً، حسابك غير مفعل حالياً.", show_alert=True)
        return

    if call.data == "client_purchases":
        history = list(transactions.find({"uid": uid, "type": "شراء"}).sort("_id", -1).limit(10))
        if not history:
            bot.answer_callback_query(call.id, "لا توجد مشتريات مسجلة في حسابك.", show_alert=True)
            return
            
        txt = "🛒 **آخر 10 مشتريات لك:**\n\n"
        for t in history:
            order_id = t.get('order_id', 'N/A')
            item_name = t.get('item_name', 'منتج')
            qty = t.get('quantity', 1)
            price = t.get('price', 0)
            date_str = t.get('date', '')
            txt += f"▪️ {date_str} | الفاتورة #{order_id} | {item_name} (x{qty}) | السعر: {price}\n"
            
    elif call.data == "client_statement":
        history = list(transactions.find({"uid": uid}).sort("_id", -1).limit(20))
        if not history:
            bot.answer_callback_query(call.id, "لا توجد حركات في حسابك.", show_alert=True)
            return
            
        txt = "🧾 **كشف حساب (آخر 20 حركة):**\n\n"
        for t in history:
            date_str = t.get('date', '')
            t_type = t.get('type', '')
            if t_type == "شراء":
                item_name = t.get('item_name', '')
                price = t.get('price', 0)
                txt += f"🔴 خصم | {date_str} | شراء {item_name} | -{price}\n"
            else:
                amount = t.get('amount', 0)
                txt += f"🟢 إضافة | {date_str} | {t_type} | +{amount}\n"
                
    bot.send_message(uid, txt, parse_mode="Markdown")
    bot.answer_callback_query(call.id)

# =======================================================
# 8. نظام الشحن الذاتي (Auto Charging)
# =======================================================

@bot.message_handler(func=lambda m: m.text == "💳 شحن")
def ask_for_card_code(msg):
    u = check_user_access(msg.chat.id)
    if not u:
        return
    
    ask_msg = bot.send_message(msg.chat.id, "يرجى إرسال كود شحن الرصيد (الكارت):")
    bot.register_next_step_handler(ask_msg, process_card_charging)

def process_card_charging(msg):
    if msg.text in MENU_BUTTONS:
        bot.send_message(msg.chat.id, "تم إلغاء عملية الشحن.")
        return
        
    uid = msg.chat.id
    u = users.find_one({"_id": uid})
    
    if not u or u.get("status") != "active":
        bot.send_message(uid, "❌ حسابك غير مفعل أو محظور.")
        return

    input_code = msg.text.strip()
    
    card = cards.find_one_and_update(
        {"code": input_code, "used": False}, 
        {"$set": {"used": True, "used_by": uid, "used_at": datetime.datetime.now()}}
    )
    
    if not card:
        attempts = u.get("failed_attempts", 0) + 1
        if attempts >= 5:
            users.update_one({"_id": uid}, {"$set": {"status": "frozen", "failed_attempts": 0}})
            bot.send_message(uid, "🚫 تم إيقاف حسابك وتجميده بسبب تكرار إدخال أكواد خاطئة. تواصل مع الإدارة.")
        else:
            users.update_one({"_id": uid}, {"$set": {"failed_attempts": attempts}})
            bot.send_message(uid, f"❌ الكود غير صحيح. متبقي لك {5 - attempts} محاولات قبل تجميد الحساب.")
        return

    card_value = float(card["value"])
    new_balance = float(u.get("balance", 0.0)) + card_value
    
    users.update_one({"_id": uid}, {"$set": {"balance": new_balance, "failed_attempts": 0}})
    
    transactions.insert_one({
        "uid": uid, 
        "user_name": u.get("name"), 
        "type": "شحن كارت", 
        "amount": card_value, 
        "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    
    bot.send_message(uid, f"✅ تم شحن حسابك بمبلغ {card_value} بنجاح.\n💰 رصيدك الجديد: {new_balance}")
    
    safe_n = safe_str(u.get('name'))
    for admin_id in get_all_admins():
        try:
            bot.send_message(
                admin_id, 
                f"💳 **عملية شحن جديدة!**\n\n👤 العميل: {safe_n}\n📱 الهاتف: `{u.get('phone')}`\n💰 المبلغ: {card_value}\n💵 الرصيد الجديد: {new_balance}",
                parse_mode="Markdown"
            )
        except Exception:
            pass

# =======================================================
# 9. نظام التسوق الديناميكي (Shopping Logic)
# =======================================================

@bot.message_handler(func=lambda m: m.text == "🛒 شراء")
def start_shopping_process(msg):
    u = check_user_access(msg.chat.id)
    if not u:
        return
    
    cats = stock.distinct("category", {"sold": False})
    if not cats:
        bot.send_message(msg.chat.id, "❌ لا توجد منتجات متوفرة حالياً في المتجر.")
        return
        
    kb = types.InlineKeyboardMarkup()
    for c in cats:
        kb.add(types.InlineKeyboardButton(c, callback_data=f"cat_{c}"))
        
    bot.send_message(msg.chat.id, "يرجى اختيار القسم:", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("cat_"))
def handle_category_selection(call):
    cat_name = call.data.split("_", 1)[1]
    subs = stock.distinct("subcategory", {"category": cat_name, "sold": False})
    
    kb = types.InlineKeyboardMarkup()
    for s in subs:
        kb.add(types.InlineKeyboardButton(s, callback_data=f"sub_{cat_name}_{s}"))
        
    bot.edit_message_text(
        f"القسم: {cat_name}\nاختر الفئة:", 
        call.message.chat.id, 
        call.message.message_id, 
        reply_markup=kb
    )

@bot.callback_query_handler(func=lambda c: c.data.startswith("sub_"))
def handle_subcategory_selection(call):
    uid = call.message.chat.id
    u = users.find_one({"_id": uid})
    if not u or u.get("status") != "active":
        return
    
    tier = u.get("tier", 1)
    parts = call.data.split("_", 2)
    cat_name = parts[1]
    sub_name = parts[2]
    
    available_items = list(stock.find({"category": cat_name, "subcategory": sub_name, "sold": False}))
    total_count = len(available_items)
    
    if total_count < 10:
        bot.answer_callback_query(call.id, "⚠️ الكمية المتبقية أقل من الحد الأدنى للشراء (10 أكواد).", show_alert=True)
        return
        
    sample_item = available_items[0]
    
    # تحديد السعر بناءً على مستوى العميل
    p1 = sample_item.get("price_1", sample_item.get("price", 0))
    p2 = sample_item.get("price_2", p1)
    p3 = sample_item.get("price_3", p1)
    
    if tier == 1:
        final_price = p1
    elif tier == 2:
        final_price = p2
    else:
        final_price = p3
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("🛒 تأكيد طلب شراء", callback_data=f"buy_{sample_item['_id']}"))
    
    info_text = (
        f"📦 المنتج: **{sample_item['name']}**\n"
        f"💰 السعر: **{final_price}**\n"
        f"📊 الكمية المتوفرة: {total_count}\n\n"
        f"⚠️ أقل كمية للطلب هي 10 ومضاعفاتها."
    )
    
    bot.send_message(uid, info_text, reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("buy_"))
def ask_for_quantity(call):
    uid = call.message.chat.id
    pid = call.data.split("_")[1]
    
    u = users.find_one({"_id": uid})
    if not u or u.get("status") != "active":
        return
    
    item = stock.find_one({"_id": ObjectId(pid), "sold": False})
    if not item:
        bot.answer_callback_query(call.id, "❌ عذراً، المنتج لم يعد متوفراً أو تم بيعه.")
        return
    
    product_name = item['name']
    available_qty = stock.count_documents({"name": product_name, "sold": False})
    
    ask_msg = bot.send_message(
        uid, 
        f"📦 المنتج: {product_name}\n"
        f"📊 المتوفر حالياً: {available_qty}\n\n"
        f"يرجى إرسال الكمية المطلوبة (يجب أن تكون 10 أو 20 أو 30 إلخ...):"
    )
    bot.register_next_step_handler(ask_msg, finalize_purchase, u, item, available_qty)

def finalize_purchase(msg, user_data, item_ref, available_qty):
    if msg.text in MENU_BUTTONS:
        bot.send_message(msg.chat.id, "تم إلغاء عملية الشراء.")
        return
        
    uid = msg.chat.id
    
    try:
        requested_qty = int(msg.text.strip())
        if requested_qty < 10 or requested_qty % 10 != 0:
            bot.send_message(uid, "❌ الكمية يجب أن تكون من مضاعفات الـ 10 (10، 20، 50...). تم إلغاء الطلب.")
            return
    except ValueError:
        bot.send_message(uid, "❌ يرجى إدخال أرقام صحيحة فقط. تم إلغاء الطلب.")
        return
        
    if requested_qty > available_qty:
        bot.send_message(uid, f"❌ الكمية المطلوبة غير متوفرة. المتاح حالياً: {available_qty}")
        return

    # حساب السعر بناءً على مستوى العميل
    tier = user_data.get("tier", 1)
    
    p1 = item_ref.get("price_1", item_ref.get("price", 0))
    p2 = item_ref.get("price_2", p1)
    p3 = item_ref.get("price_3", p1)
    
    if tier == 1:
        unit_price = p1
    elif tier == 2:
        unit_price = p2
    else:
        unit_price = p3
        
    total_cost = requested_qty * unit_price
    
    # فحص الرصيد
    user_balance = float(user_data.get("balance", 0.0))
    if user_balance < total_cost:
        bot.send_message(uid, f"❌ رصيدك غير كافي لإتمام العملية.\nالمطلوب: {total_cost}\nرصيدك الحالي: {user_balance}")
        return

    # سحب الأكواد من قاعدة البيانات
    product_name = item_ref['name']
    batch = list(stock.find({"name": product_name, "sold": False}).limit(requested_qty))
    
    if len(batch) < requested_qty:
        bot.send_message(uid, "❌ حدث خطأ في توافر الأكواد بسبب سحب آخر. يرجى المحاولة مرة أخرى.")
        return
        
    ids_to_sell = []
    for doc in batch:
        ids_to_sell.append(doc['_id'])
        
    update_res = stock.update_many(
        {"_id": {"$in": ids_to_sell}, "sold": False}, 
        {"$set": {"sold": True, "buyer_id": uid, "order_date": datetime.datetime.now()}}
    )
    
    if update_res.modified_count != requested_qty:
        # إرجاع الأكواد في حال الفشل
        stock.update_many({"_id": {"$in": ids_to_sell}}, {"$set": {"sold": False}})
        bot.send_message(uid, "❌ حدث تضارب في عملية البيع، يرجى المحاولة مرة أخرى.")
        return

    # خصم الرصيد
    users.update_one({"_id": uid}, {"$inc": {"balance": -total_cost}})
    
    # تسجيل الفاتورة في المعاملات
    order_id = get_next_order_id()
    dt_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    transactions.insert_one({
        "order_id": order_id, 
        "uid": uid, 
        "user_name": user_data.get("name"), 
        "phone": user_data.get("phone"), 
        "type": "شراء", 
        "item_name": product_name, 
        "quantity": requested_qty, 
        "price": total_cost, 
        "date": dt_str
    })
    
    # إرسال إلى Google Sheets إن وجد
    if SHEET_WEBHOOK_URL and SHEET_WEBHOOK_URL.startswith("http"):
        try:
            payload = {
                "order_id": order_id,
                "date": dt_str,
                "phone": user_data.get('phone'),
                "item_name": f"{product_name} (x{requested_qty})",
                "price": total_cost
            }
            requests.post(SHEET_WEBHOOK_URL, json=payload, timeout=3)
        except Exception:
            pass
            
    bot.send_message(uid, f"✅ تم الشراء بنجاح!\n🧾 رقم الفاتورة: #{order_id}\n💰 القيمة المخصومة: {total_cost}\n\nجاري إعداد ملف الأكواد الخاص بك...")
    
    # إعداد الفاتورة بصيغة إكسيل ديناميكي
    purchased_items_data = []
    for d in batch:
        purchased_items_data.append({
            'code': d['code'], 
            'serial': d.get('serial', ''), 
            'pin': d.get('pin', ''), 
            'op_code': d.get('op_code', '')
        })
        
    file_stream = generate_customer_excel_file(purchased_items_data, order_id, dt_str, product_name)
    bot.send_document(uid, file_stream, caption=f"📁 فاتورة الأكواد | رقم #{order_id}")
    
    # إشعار الإدارة والمشرفين
    safe_n = safe_str(user_data.get('name'))
    admin_msg = (
        f"🛒 **عملية شراء جديدة!**\n\n"
        f"الفاتورة: #{order_id}\n"
        f"👤 العميل: {safe_n}\n"
        f"📱 الهاتف: {user_data.get('phone')}\n"
        f"🎚️ المستوى: {tier}\n"
        f"📦 المنتج: {product_name}\n"
        f"🔢 الكمية: {requested_qty}\n"
        f"💰 المبلغ المدفوع: {total_cost}"
    )
    
    for admin_id in get_all_admins():
        try:
            file_stream.seek(0)
            bot.send_message(admin_id, admin_msg)
            bot.send_document(admin_id, file_stream, caption=f"📁 نسخة فاتورة #{order_id}")
        except Exception:
            pass

    # فحص المخزون المتبقي
    remaining_stock = stock.count_documents({"name": product_name, "sold": False})
    if remaining_stock <= 30:
        for admin_id in get_all_admins():
            try:
                bot.send_message(
                    admin_id, 
                    f"⚠️ **تنبيه انخفاض المخزون** ⚠️\n\nالمنتج: `{product_name}`\nالمتبقي: **{remaining_stock}** كود فقط.", 
                    parse_mode="Markdown"
                )
            except Exception:
                pass

# =======================================================
# 10. لوحة تحكم الإدارة (Admin Dashboard)
# =======================================================

@bot.message_handler(commands=['admin'])
def show_admin_dashboard(msg):
    if not is_admin(msg.chat.id):
        return
    bot.send_message(msg.chat.id, "👑 مرحباً بك في لوحة تحكم الإدارة العليا.", reply_markup=admin_menu())

@bot.message_handler(func=lambda m: m.text == "🏪 العودة للمتجر")
def return_to_client_mode(msg):
    if not is_admin(msg.chat.id):
        return
    bot.send_message(msg.chat.id, "🔄 تم تحويلك لوضع العميل.", reply_markup=menu())

# --- أ. قائمة المستخدمين ---
@bot.message_handler(func=lambda m: m.text == "👥 المستخدمين")
def admin_users_list(msg):
    if not is_admin(msg.chat.id):
        return
        
    try:
        text = "👥 **قائمة آخر 30 مستخدم مسجل في النظام:**\n\n"
        users_list_cursor = users.find().sort("join", -1).limit(30)
        
        for u in users_list_cursor:
            stat = "✅" if u.get('status') == 'active' else ("🚫" if u.get('status') == 'blocked' else "❄️")
            safe_n = safe_str(u.get('name'))
            phone = u.get('phone', 'بدون')
            balance = u.get('balance', 0)
            
            text += f"📛 {safe_n} | 📱 `{phone}` | 💰 {balance} | {stat}\n"
            
        bot.send_message(msg.chat.id, text, parse_mode="Markdown")
    except Exception as e:
        bot.send_message(msg.chat.id, f"❌ حدث خطأ أثناء جلب قائمة المستخدمين:\n{e}")

# --- ب. سجل الفواتير ---
@bot.message_handler(func=lambda m: m.text == "🧾 سجل الفواتير")
def admin_invoices_log(msg):
    if not is_admin(msg.chat.id):
        return
        
    try:
        history = list(transactions.find({"type": "شراء", "order_id": {"$exists": True}}).sort("_id", -1).limit(40))
        if not history:
            bot.send_message(msg.chat.id, "لا توجد فواتير مبيعات مسجلة حتى الآن.")
            return
            
        text = "🧾 **سجل آخر 40 فاتورة مبيعات:**\n\n"
        for t in history:
            order_id = t.get('order_id', '-')
            safe_n = safe_str(t.get('user_name'))
            item_name = t.get('item_name', '-')
            qty = t.get('quantity', 1)
            price = t.get('price', 0)
            
            text += f"▪️ #{order_id} | 👤 {safe_n} | 🛒 {item_name} (x{qty}) | 💰 {price}\n"
            
        bot.send_message(msg.chat.id, text, parse_mode="Markdown")
    except Exception as e:
        bot.send_message(msg.chat.id, f"❌ حدث خطأ:\n{e}")

# --- ج. إدارة العميل (تفعيل، حظر، مستوى، تقرير) ---
@bot.message_handler(func=lambda m: m.text == "⚙️ إدارة عميل")
def admin_manage_customer(msg):
    if not is_admin(msg.chat.id):
        return
        
    ask_msg = bot.send_message(msg.chat.id, "يرجى إرسال رقم هاتف العميل للبحث عنه:")
    bot.register_next_step_handler(ask_msg, admin_find_and_render_customer)

def admin_find_and_render_customer(msg):
    if msg.text in MENU_BUTTONS:
        return
        
    u = find_customer(msg.text)
    if not u:
        bot.send_message(msg.chat.id, "❌ لم يتم العثور على العميل.")
        return
        
    admin_render_customer_card(u["_id"], msg.chat.id)

def admin_render_customer_card(uid, chat_id, message_id=None):
    u = users.find_one({"_id": uid})
    if not u:
        return
    
    tier = u.get("tier", 1)
    if tier == 1:
        tier_label = "مستوى 1 🥉"
    elif tier == 2:
        tier_label = "مستوى 2 🥈"
    else:
        tier_label = "مستوى 3 🥇"
        
    stat = u.get("status")
    if stat == "blocked":
        stat_label = "محظور 🚫"
    elif stat == "active":
        stat_label = "نشط ✅"
    else:
        stat_label = "مجمد ❄️"
    
    safe_n = safe_str(u.get('name'))
    info = (
        f"👤 **ملف بيانات العميل:**\n\n"
        f"📛 الاسم: {safe_n}\n"
        f"🆔 ID: `{uid}`\n"
        f"📱 الهاتف: `{u.get('phone')}`\n"
        f"💰 الرصيد: **{u.get('balance',0)}**\n"
        f"🎚️ المستوى الحالي: **{tier_label}**\n"
        f"حالة الحساب: **{stat_label}**"
    )
    
    kb = types.InlineKeyboardMarkup(row_width=2)
    
    if stat == "active":
        kb.add(
            types.InlineKeyboardButton("❄️ تجميد الحساب", callback_data=f"adm_freeze_{uid}"), 
            types.InlineKeyboardButton("🚫 حظر و طرد", callback_data=f"adm_block_{uid}")
        )
    elif stat == "frozen":
        kb.add(
            types.InlineKeyboardButton("✅ تفعيل الحساب", callback_data=f"adm_activate_{uid}"), 
            types.InlineKeyboardButton("🚫 حظر و طرد", callback_data=f"adm_block_{uid}")
        )
    elif stat == "blocked":
        kb.add(
            types.InlineKeyboardButton("✅ فك الحظر", callback_data=f"adm_activate_{uid}")
        )
        
    kb.add(
        types.InlineKeyboardButton("🎚️ تغيير مستوى الأسعار", callback_data=f"adm_chgtier_{uid}"), 
        types.InlineKeyboardButton("📊 تقرير عمليات العميل", callback_data=f"adm_report_{uid}")
    )
    
    if message_id:
        try:
            bot.edit_message_text(info, chat_id, message_id, reply_markup=kb, parse_mode="Markdown")
        except Exception:
            pass
    else:
        bot.send_message(chat_id, info, reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("adm_"))
def handle_admin_customer_actions(call):
    parts = call.data.split("_")
    action = parts[1]
    uid = int(parts[2])
    
    if action == "report":
        history = list(transactions.find({"uid": uid}).sort("_id", -1).limit(15))
        if not history:
            bot.answer_callback_query(call.id, "لا توجد عمليات لهذا العميل.")
            return
            
        txt = f"📊 **آخر عمليات العميل `{uid}`:**\n\n"
        for t in history:
            date_str = t.get('date', '')
            t_type = t.get('type', '')
            price = t.get('price', t.get('amount', 0))
            txt += f"▪️ {date_str} | {t_type} | {price}\n"
            
        bot.send_message(call.message.chat.id, txt, parse_mode="Markdown")
        bot.answer_callback_query(call.id)
        return

    if action == "chgtier":
        kb = types.InlineKeyboardMarkup(row_width=3)
        kb.add(
            types.InlineKeyboardButton("المستوى 1 🥉", callback_data=f"set_tier_1_{uid}"),
            types.InlineKeyboardButton("المستوى 2 🥈", callback_data=f"set_tier_2_{uid}"),
            types.InlineKeyboardButton("المستوى 3 🥇", callback_data=f"set_tier_3_{uid}")
        )
        kb.add(types.InlineKeyboardButton("🔙 رجوع للخلف", callback_data=f"adm_back_{uid}"))
        
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=kb)
        return

    if action == "freeze":
        users.update_one({"_id": uid}, {"$set": {"status": "frozen"}})
    elif action == "activate":
        users.update_one({"_id": uid}, {"$set": {"status": "active", "failed_attempts": 0}})
    elif action == "block":
        users.update_one({"_id": uid}, {"$set": {"status": "blocked"}})
    elif action == "back":
        pass
    
    bot.answer_callback_query(call.id, "تم تحديث بيانات العميل بنجاح.")
    admin_render_customer_card(uid, call.message.chat.id, call.message.message_id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("set_tier_"))
def handle_tier_setting(call):
    parts = call.data.split("_")
    level = int(parts[2])
    uid = int(parts[3])
    
    users.update_one({"_id": uid}, {"$set": {"tier": level}})
    bot.answer_callback_query(call.id, f"تم تغيير مستوى العميل إلى {level}")
    admin_render_customer_card(uid, call.message.chat.id, call.message.message_id)

# --- د. إضافة منتجات (القالب و اليدوي) ---
@bot.message_handler(func=lambda m: m.text == "➕ منتج")
def admin_add_product(msg):
    if not is_admin(msg.chat.id):
        return
        
    txt = (
        "➕ **إضافة منتجات للمخزن:**\n\n"
        "📁 **الطريقة الأولى (الرفع الشامل):** قم بتحميل قالب الإكسيل المرفق، املأه بالأكواد وأعد رفعه.\n"
        "✍️ **الطريقة الثانية (الإضافة اليدوية):** أرسل البيانات بالتنسيق التالي:\n"
        "`القسم:الفئة:الاسم:السعر1:السعر2:السعر3`\n"
        "*(إذا أردت سعراً واحداً أرسل السعر الأول فقط وسيتم تعميمه)*"
    )
    
    try:
        template_file = generate_products_template()
        ask_msg = bot.send_document(msg.chat.id, template_file, caption=txt, parse_mode="Markdown")
        bot.register_next_step_handler(ask_msg, admin_process_product_input)
    except Exception as e:
        bot.send_message(msg.chat.id, f"❌ حدث خطأ في توليد القالب: {e}")

def admin_process_product_input(msg):
    if msg.text in MENU_BUTTONS:
        return
    
    if msg.document:
        if not msg.document.file_name.endswith('.xlsx'):
            bot.send_message(msg.chat.id, "❌ يرجى رفع ملف إكسيل بصيغة .xlsx فقط.")
            return
            
        bot.send_message(msg.chat.id, "⏳ جاري قراءة البيانات من الملف...")
        try:
            file_info = bot.get_file(msg.document.file_id)
            downloaded = bot.download_file(file_info.file_path)
            wb = openpyxl.load_workbook(io.BytesIO(downloaded))
            ws = wb.active
            
            batch_data = []
            errors_count = 0
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or not row[0]:
                    continue
                    
                try:
                    category = str(row[0]).strip()
                    subcategory = str(row[1]).strip()
                    name = str(row[2]).strip()
                    
                    price_1 = float(row[3])
                    price_2 = float(row[4]) if len(row) > 4 and row[4] else price_1
                    price_3 = float(row[5]) if len(row) > 5 and row[5] else price_1
                    
                    code_val = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    serial_val = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                    pin_val = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                    opcode_val = str(row[9]).strip() if len(row) > 9 and row[9] else ""
                    
                    if not category or not subcategory or not name or not code_val:
                        errors_count += 1
                        continue
                        
                    batch_data.append({
                        "category": category, 
                        "subcategory": subcategory, 
                        "name": name,
                        "price_1": price_1, 
                        "price_2": price_2, 
                        "price_3": price_3,
                        "code": code_val, 
                        "serial": serial_val, 
                        "pin": pin_val, 
                        "op_code": opcode_val
                    })
                except Exception:
                    errors_count += 1
                    
            if errors_count > 0:
                bot.send_message(msg.chat.id, f"⚠️ تم تخطي {errors_count} صف لوجود بيانات ناقصة أو غير صحيحة.")
                
            # تحويل البيانات لنظام الرادار
            process_radar_logic(msg.chat.id, batch_data)
            
        except Exception as e:
            bot.send_message(msg.chat.id, f"❌ حدث خطأ أثناء معالجة الملف:\n{e}")
        
    elif msg.text and ":" in msg.text:
        try:
            parts = msg.text.split(":")
            cat = parts[0].strip()
            sub = parts[1].strip()
            name = parts[2].strip()
            
            p1 = float(parts[3].strip())
            p2 = float(parts[4].strip()) if len(parts) > 4 else p1
            p3 = float(parts[5].strip()) if len(parts) > 5 else p1
            
            product_info = {
                "cat": cat, 
                "sub": sub, 
                "name": name, 
                "price_1": p1, 
                "price_2": p2, 
                "price_3": p3
            }
            
            temp_admin_data[msg.chat.id] = {"info": product_info}
            ask_msg = bot.send_message(
                msg.chat.id, 
                "✅ تم حفظ بيانات المنتج والأسعار.\n\nالآن أرسل الأكواد كـ **رسالة نصية** بالتنسيق التالي:\n`الكود:التسلسلي:PIN:أوبريشن`\n\n*(السطر الواحد يمثل كوداً واحداً)*"
            )
            bot.register_next_step_handler(ask_msg, admin_process_manual_codes)
        except Exception:
            bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح. تأكد من إدخال البيانات مفصولة بنقطتين (:).")

def admin_process_manual_codes(msg):
    if msg.text in MENU_BUTTONS:
        return
        
    p_info = temp_admin_data.get(msg.chat.id, {}).get("info")
    if not p_info:
        bot.send_message(msg.chat.id, "❌ انتهت جلسة الإضافة. حاول من جديد.")
        return
    
    batch_data = []
    if msg.text:
        lines = msg.text.split("\n")
        for line in lines:
            if not line.strip():
                continue
                
            pts = line.split(":")
            code_val = pts[0].strip()
            serial_val = pts[1].strip() if len(pts) > 1 else ""
            pin_val = pts[2].strip() if len(pts) > 2 else ""
            op_val = pts[3].strip() if len(pts) > 3 else ""
            
            batch_data.append({
                "code": code_val, 
                "serial": serial_val,
                "pin": pin_val, 
                "op_code": op_val
            })
            
    # تحويل الأكواد لنظام الرادار
    process_radar_logic(msg.chat.id, batch_data, p_info)
    
    # تنظيف الجلسة
    if msg.chat.id in temp_admin_data:
        del temp_admin_data[msg.chat.id]

# --- هـ. تحديث الأسعار المجمع ---
@bot.message_handler(func=lambda m: m.text == "💵 أسعار المستويات")
def admin_bulk_price_edit(msg):
    if not is_admin(msg.chat.id):
        return
        
    # تجميع جميع المنتجات الفريدة في المخزن
    pipeline = [
        {"$match": {"sold": False}}, 
        {"$group": {
            "_id": "$name", 
            "cat": {"$first": "$category"}, 
            "sub": {"$first": "$subcategory"}, 
            "p1": {"$first": "$price_1"}, 
            "p2": {"$first": "$price_2"}, 
            "p3": {"$first": "$price_3"}
        }}
    ]
    prods = list(stock.aggregate(pipeline))
    
    if not prods:
        bot.send_message(msg.chat.id, "❌ لا توجد منتجات متاحة للبيع في المخزن لتعديل أسعارها.")
        return
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تحديث الأسعار"
    
    headers = ["الاسم (لا تقم بتعديله أبداً)", "القسم", "الفئة", "السعر للمستوى 1", "السعر للمستوى 2", "السعر للمستوى 3"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        
    for p in prods:
        name = p["_id"]
        cat = p.get("cat", "")
        sub = p.get("sub", "")
        p1 = p.get("p1", 0)
        p2 = p.get("p2", 0)
        p3 = p.get("p3", 0)
        
        ws.append([name, cat, sub, p1, p2, p3])
        
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = max_length + 4

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    stream.name = "Prices_Update_List.xlsx"
    
    ask_msg = bot.send_document(
        msg.chat.id, 
        stream, 
        caption="📁 **أداة تعديل الأسعار الشاملة**\n\nقم بتحميل الملف، عدل الأسعار في الأعمدة الثلاثة الأخيرة حسب رغبتك، ثم أعد إرسال الملف هنا ليتم التحديث فوراً."
    )
    bot.register_next_step_handler(ask_msg, admin_finalize_price_update)

def admin_finalize_price_update(msg):
    if msg.text in MENU_BUTTONS:
        return
        
    if not msg.document or not msg.document.file_name.endswith('.xlsx'):
        bot.send_message(msg.chat.id, "❌ يرجى رفع ملف الإكسيل الذي قمت بتعديله بصيغة .xlsx")
        return
        
    bot.send_message(msg.chat.id, "⏳ جاري قراءة الملف وتحديث الأسعار في قاعدة البيانات...")
    try:
        file_info = bot.get_file(msg.document.file_id)
        downloaded = bot.download_file(file_info.file_path)
        wb = openpyxl.load_workbook(io.BytesIO(downloaded))
        ws = wb.active
        
        updated_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]:
                continue
                
            name = str(row[0]).strip()
            try:
                p1 = float(row[3]) if row[3] is not None else 0
                p2 = float(row[4]) if row[4] is not None else p1
                p3 = float(row[5]) if row[5] is not None else p1
                
                res = stock.update_many(
                    {"name": name, "sold": False}, 
                    {"$set": {"price_1": p1, "price_2": p2, "price_3": p3}}
                )
                if res.modified_count > 0:
                    updated_count += 1
            except Exception:
                continue
                
        bot.send_message(msg.chat.id, f"✅ تمت العملية بنجاح. تم تحديث أسعار `{updated_count}` نوع من المنتجات لجميع المستويات.")
    except Exception as e:
        bot.send_message(msg.chat.id, f"❌ حدث خطأ أثناء المعالجة:\n{e}")

# --- و. إدارة المخزون الداخلي (تعديل فردي ومسح) ---
@bot.message_handler(func=lambda m: m.text == "📦 إدارة المخزون")
def manage_stock_cmd(msg):
    if not is_admin(msg.chat.id):
        return
        
    names = stock.distinct("name", {"sold": False})
    if not names:
        bot.send_message(msg.chat.id, "❌ المخزن فارغ حالياً.")
        return
        
    text = "📦 **المنتجات المتوفرة حالياً في المخزن:**\n\n"
    for n in names:
        count = stock.count_documents({"name": n, "sold": False})
        text += f"▪️ `{n}` (الكمية: {count})\n"
        
    text += "\n👉 أرسل **اسم المنتج** (انسخه من القائمة بالأعلى) لإدارته:"
    ask_msg = bot.send_message(msg.chat.id, text, parse_mode="Markdown")
    bot.register_next_step_handler(ask_msg, show_stock_item_panel)

def show_stock_item_panel(msg):
    if msg.text in MENU_BUTTONS:
        return
        
    name = msg.text.strip()
    count = stock.count_documents({"name": name, "sold": False})
    item = stock.find_one({"name": name, "sold": False})
    
    if not item:
        bot.send_message(msg.chat.id, "❌ المنتج غير موجود في المخزن.")
        return
        
    p1 = item.get("price_1", item.get("price", 0))
    p2 = item.get("price_2", p1)
    p3 = item.get("price_3", p1)
        
    text = (
        f"📦 المنتج: `{name}`\n"
        f"💰 الأسعار الحالية:\n"
        f"مستوى 1: {p1} | مستوى 2: {p2} | مستوى 3: {p3}\n"
        f"📊 الكمية المتوفرة: {count}"
    )
    
    if msg.chat.id not in temp_admin_data:
        temp_admin_data[msg.chat.id] = {}
        
    temp_admin_data[msg.chat.id]["mng_item_name"] = name
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("💰 تعديل السعر (فردي)", callback_data="stk_price"))
    kb.add(types.InlineKeyboardButton("➖ حذف كمية معينة", callback_data="stk_delqty"))
    kb.add(types.InlineKeyboardButton("❌ حذف المنتج بالكامل", callback_data="stk_delall"))
    
    bot.send_message(msg.chat.id, text, reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("stk_"))
def handle_stock_actions(call):
    action = call.data.split("_")[1]
    data = temp_admin_data.get(call.message.chat.id)
    
    if not data or "mng_item_name" not in data:
        bot.answer_callback_query(call.id, "❌ انتهت الجلسة.", show_alert=True)
        return
        
    name = data["mng_item_name"]
    
    if action == "price":
        ask_msg = bot.send_message(call.message.chat.id, f"أرسل الأسعار الجديدة للمنتج `{name}` بالتنسيق:\nسعر1:سعر2:سعر3", parse_mode="Markdown")
        bot.register_next_step_handler(ask_msg, admin_update_stock_price, name)
        bot.answer_callback_query(call.id)
        
    elif action == "delqty":
        ask_msg = bot.send_message(call.message.chat.id, f"أرسل عدد الأكواد المراد مسحها من منتج `{name}`:")
        bot.register_next_step_handler(ask_msg, admin_delete_stock_qty, name)
        bot.answer_callback_query(call.id)
        
    elif action == "delall":
        res = stock.delete_many({"name": name, "sold": False})
        bot.answer_callback_query(call.id, f"تم مسح {res.deleted_count} كود.")
        bot.edit_message_text(f"✅ تم حذف المنتج `{name}` بالكامل.", call.message.chat.id, call.message.message_id, parse_mode="Markdown")

def admin_update_stock_price(msg, name):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        p1 = float(parts[0].strip())
        p2 = float(parts[1].strip()) if len(parts) > 1 else p1
        p3 = float(parts[2].strip()) if len(parts) > 2 else p1
        
        res = stock.update_many({"name": name, "sold": False}, {"$set": {"price_1": p1, "price_2": p2, "price_3": p3}})
        bot.send_message(msg.chat.id, f"✅ تم تحديث الأسعار لـ {res.modified_count} كود من هذا المنتج.")
    except Exception:
        bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

def admin_delete_stock_qty(msg, name):
    if msg.text in MENU_BUTTONS: return
    try:
        qty_to_delete = int(msg.text.strip())
        docs_to_delete = list(stock.find({"name": name, "sold": False}).limit(qty_to_delete))
        if not docs_to_delete:
            bot.send_message(msg.chat.id, "❌ لا يوجد رصيد متاح للحذف.")
            return
            
        ids = [d['_id'] for d in docs_to_delete]
        res = stock.delete_many({"_id": {"$in": ids}})
        bot.send_message(msg.chat.id, f"✅ تم حذف {res.deleted_count} كود بنجاح.")
    except Exception:
        bot.send_message(msg.chat.id, "❌ يجب إرسال أرقام صحيحة.")

# --- ز. استخراج تقارير الإكسيل مع فلترة التاريخ ---
@bot.message_handler(func=lambda m: m.text == "📊 تقارير إكسيل")
def admin_excel_reports(msg):
    if not is_admin(msg.chat.id):
        return
        
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("📄 تقرير شامل للمبيعات", callback_data="rep_all"),
           types.InlineKeyboardButton("👤 تقرير مخصص لعميل", callback_data="rep_single"))
           
    bot.send_message(msg.chat.id, "الرجاء اختيار نوع التقرير الذي تريد استخراجه:", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("rep_"))
def admin_report_dates(call):
    rep_type = call.data.split("_")[1]
    
    if call.message.chat.id not in temp_admin_data:
        temp_admin_data[call.message.chat.id] = {}
        
    temp_admin_data[call.message.chat.id]["rep_type"] = rep_type
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("🕒 كل الأوقات والتواريخ", callback_data="dt_all"),
           types.InlineKeyboardButton("📅 تحديد فترة (من-إلى)", callback_data="dt_custom"))
           
    bot.edit_message_text("هل تريد استخراج التقرير لكل الأوقات، أم تحديد فترة زمنية معينة؟", call.message.chat.id, call.message.message_id, reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("dt_"))
def admin_report_finalize(call):
    uid = call.message.chat.id
    mode = call.data.split("_")[1]
    
    rep_type = temp_admin_data.get(uid, {}).get("rep_type")
    if not rep_type:
        bot.answer_callback_query(call.id, "انتهت الجلسة، حاول مرة أخرى.", show_alert=True)
        return
        
    bot.answer_callback_query(call.id)
    
    if mode == "all":
        admin_execute_report(uid, rep_type, None, None)
    else:
        ask_msg = bot.send_message(
            uid, 
            "أرسل التاريخ (من) و (إلى) بالتنسيق التالي بدقة:\n`YYYY-MM-DD:YYYY-MM-DD`\n\nمثال:\n`2026-03-01:2026-04-10`", 
            parse_mode="Markdown"
        )
        bot.register_next_step_handler(ask_msg, admin_process_date_input, rep_type)

def admin_process_date_input(msg, rep_type):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        start_dt = parts[0].strip()
        end_dt = parts[1].strip()
        admin_execute_report(msg.chat.id, rep_type, start_dt, end_dt)
    except Exception:
        bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح. قم بإعادة الطلب وتأكد من الفاصل (:).")

def admin_execute_report(uid, rep_type, start_date, end_date):
    filtr = {}
    if start_date and end_date:
        filtr = {"date": {"$gte": start_date, "$lte": end_date + " 23:59"}}
    
    if rep_type == "all":
        bot.send_message(uid, "⏳ جاري تجميع بيانات المبيعات وتجهيز ملف الإكسيل...")
        
        data = list(transactions.find(filtr).sort("_id", -1))
        if not data:
            bot.send_message(uid, "❌ لا توجد عمليات مسجلة في النظام خلال هذه الفترة.")
            return
            
        summary = {}
        for t in data:
            phone = t.get("phone", "بدون رقم")
            if t.get("type") == "شراء":
                if phone not in summary:
                    summary[phone] = {"name": t.get("user_name", "غير مسجل"), "spent": 0}
                summary[phone]["spent"] += t.get("price", 0)
                
        file_stream = generate_admin_report_excel("all", data, summary)
        bot.send_document(uid, file_stream, caption="✅ تم تجهيز التقرير الشامل.")
        
    else:
        ask_msg = bot.send_message(uid, "يرجى إرسال رقم الهاتف للعميل المطلوب استخراج تقريره:")
        bot.register_next_step_handler(ask_msg, admin_execute_single_report, filtr)

def admin_execute_single_report(msg, filtr):
    if msg.text in MENU_BUTTONS: return
    
    u = find_customer(msg.text)
    if not u:
        bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
        return
        
    query = {"uid": u["_id"]}
    query.update(filtr)
    
    data = list(transactions.find(query).sort("_id", -1))
    if not data:
        bot.send_message(msg.chat.id, "❌ لا توجد عمليات مسجلة لهذا العميل في هذه الفترة.")
        return
        
    file_stream = generate_admin_report_excel("single", data)
    file_stream.name = f"Report_{u.get('phone', 'Client')}.xlsx"
    bot.send_document(msg.chat.id, file_stream, caption=f"✅ تقرير مخصص للعميل: {safe_str(u.get('name'))}")

# --- ح. إدارة الرصيد وكروت الشحن ---
@bot.message_handler(func=lambda m: m.text == "💰 ضبط الرصيد")
def admin_set_balance(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل التعديل بالتنسيق (رقم_الهاتف:الرصيد_الجديد):\nمثال: `0940719000:500`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_set_balance)

def execute_set_balance(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        u = find_customer(parts[0])
        if u:
            new_bal = float(parts[1].strip())
            users.update_one({"_id": u["_id"]}, {"$set": {"balance": new_bal}})
            transactions.insert_one({"uid": u["_id"], "user_name": u.get("name"), "type": "ضبط رصيد من الإدارة", "amount": new_bal, "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
            bot.send_message(msg.chat.id, f"✅ تم ضبط رصيد العميل ليصبح {new_bal}")
            try: bot.send_message(u["_id"], f"⚙️ إشعار من الإدارة: تم تحديث رصيدك ليصبح {new_bal}")
            except: pass
        else:
            bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    except Exception: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

@bot.message_handler(func=lambda m: m.text == "💳 شحن يدوي")
def admin_direct_charge(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل الشحن بالتنسيق (رقم_الهاتف:قيمة_الإضافة):\nمثال: `0940719000:50`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_direct_charge)

def execute_direct_charge(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        u = find_customer(parts[0])
        if u:
            amt = float(parts[1].strip())
            users.update_one({"_id": u["_id"]}, {"$inc": {"balance": amt}})
            transactions.insert_one({"uid": u["_id"], "user_name": u.get("name"), "type": "إضافة رصيد مباشر", "amount": amt, "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
            bot.send_message(msg.chat.id, f"✅ تمت الإضافة بنجاح.")
            try: bot.send_message(u["_id"], f"🎁 إشعار من الإدارة: تم إيداع {amt} في حسابك.")
            except: pass
        else:
            bot.send_message(msg.chat.id, "❌ العميل غير موجود.")
    except Exception: bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

@bot.message_handler(func=lambda m: m.text == "🎫 توليد")
def admin_generate_cards(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل العدد والقيمة بالتنسيق (العدد:القيمة)\nمثال لتوليد 10 كروت فئة 50:\n`10:50`", parse_mode="Markdown")
    bot.register_next_step_handler(ask, execute_generate_cards)

def execute_generate_cards(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        parts = msg.text.split(":")
        count = int(parts[0].strip())
        val = float(parts[1].strip())
        
        arr = []
        result_text = f"✅ تم توليد {count} كروت شحن فئة {val}:\n\n"
        
        for _ in range(count):
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=12))
            arr.append({"code": code, "value": val, "used": False})
            result_text += f"`{code}`\n"
            
        cards.insert_many(arr)
        bot.send_message(msg.chat.id, result_text, parse_mode="Markdown")
    except Exception:
        bot.send_message(msg.chat.id, "❌ التنسيق غير صحيح.")

# =======================================================
# 11. الأوامر المخفية (Hidden Secret Commands)
# =======================================================

@bot.message_handler(commands=['FRP', 'frp'])
def frp_reset_command(msg):
    if msg.chat.id != OWNER_ID: return
    ask = bot.send_message(
        msg.chat.id, 
        "⚠️ **تحذير: أمر فورمات المصنع الشامل** ⚠️\n\nأنت على وشك مسح (المستخدمين، المخزون، الفواتير، الأرصدة، المدراء، الكروت).\n\nأرسل هذه الجملة حرفياً للتأكيد:\n`تأكيد الحذف النهائي`", 
        parse_mode="Markdown"
    )
    bot.register_next_step_handler(ask, execute_frp_wipe)

def execute_frp_wipe(msg):
    if msg.text == "تأكيد الحذف النهائي":
        users.delete_many({})
        stock.delete_many({})
        transactions.delete_many({})
        admins_db.delete_many({})
        counters.delete_many({})
        cards.delete_many({})
        bot.send_message(msg.chat.id, "✅ تم تصفير النظام وعودته لحالة المصنع بالكامل.")
    else:
        bot.send_message(msg.chat.id, "❌ تم إلغاء عملية المسح.")

@bot.message_handler(commands=['ADD', 'add'])
def add_sub_admin_command(msg):
    if msg.chat.id != OWNER_ID: return
    ask = bot.send_message(msg.chat.id, "أرسل رقم الـ ID الخاص بالشخص لترقيته إلى مدير:")
    bot.register_next_step_handler(ask, execute_add_admin)

def execute_add_admin(msg):
    if msg.text in MENU_BUTTONS: return
    try:
        new_admin_id = int(msg.text.strip())
        admins_db.update_one(
            {"_id": new_admin_id}, 
            {"$set": {"added_by": msg.chat.id, "join_date": datetime.datetime.now()}}, 
            upsert=True
        )
        bot.send_message(msg.chat.id, f"✅ تمت ترقية `{new_admin_id}` بنجاح.", parse_mode="Markdown")
        try: bot.send_message(new_admin_id, "🎉 تهانينا! تمت ترقيتك لمشرف في النظام. أرسل /admin للدخول للوحة.")
        except: pass
    except Exception:
        bot.send_message(msg.chat.id, "❌ يرجى إرسال ID صحيح (أرقام فقط).")

@bot.message_handler(commands=['Block', 'block'])
def block_user_command(msg):
    if not is_admin(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "أرسل رقم الهاتف المراد حظره وطرد صاحبه من النظام:")
    bot.register_next_step_handler(ask, execute_block)

def execute_block(msg):
    if msg.text in MENU_BUTTONS: return
    u = find_customer(msg.text)
    if u:
        users.update_one({"_id": u["_id"]}, {"$set": {"status": "blocked"}})
        bot.send_message(msg.chat.id, "✅ تم الحظر.")
        try: bot.send_message(u["_id"], "تم حظرك من قبل الادارة وللاستفسار تواصل معنا", reply_markup=types.ReplyKeyboardRemove())
        except: pass
    else:
        bot.send_message(msg.chat.id, "❌ العميل غير موجود.")

# =======================================================
# 12. محرك تشغيل السيرفر (Render Background Setup)
# =======================================================

app = Flask(__name__)

@app.route('/')
def health_check_route():
    return "<h1>شركة الأهرام للإتصالات - سيرفر البوت يعمل بنجاح</h1>"

def bot_polling_worker():
    """هذه الدالة تعزل عمل البوت عن الويب سيرفر لمنع إيقافه من المنصة"""
    bot.remove_webhook()
    print("🚀 STARTED: Telegram Polling Initialized...")
    
    while True:
        try:
            # تشغيل البوت مع إعدادات تجنب فقدان الاتصال
            bot.infinity_polling(timeout=30, long_polling_timeout=15)
        except Exception as e:
            print(f"⚠️ POLLING ERROR RESTARTING: {e}")
            time.sleep(5)

if __name__ == "__main__":
    # تشغيل البوت في خيط (Thread) خلفي لا يتداخل مع الفلاسك
    polling_thread = threading.Thread(target=bot_polling_worker)
    polling_thread.daemon = True
    polling_thread.start()
    
    # تشغيل سيرفر الويب لربط الـ Port الخاص بـ Render
    port_number = int(os.environ.get("PORT", 8080))
    print(f"🌍 WEB SERVER: Listening on Port {port_number}")
    app.run(host="0.0.0.0", port=port_number)
