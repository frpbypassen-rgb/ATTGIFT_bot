import telebot
from telebot import types
from pymongo import MongoClient, ReturnDocument
from bson import ObjectId
import datetime
import certifi
import os
from flask import Flask
import threading
import time
import requests 
import io
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# إعدادات بوت العملاء
# ==========================================
CUSTOMER_BOT_TOKEN = "8769145956:AAEKIAKJ2sGn9HFu_-M8diyND1J754fp_Wc"
ADMIN_IDS = [1262656649] # أرقام الإدارة لتلقي الإشعارات

MONGO_URI = "mongodb+srv://frpbypassen_db_user:LpovkVYkrNU7qePp@attgift.rdamxpj.mongodb.net/?retryWrites=true&w=majority&appName=ATTGIFT"
client = MongoClient(MONGO_URI, tlsCAFile=certifi.where(), maxPoolSize=50, connectTimeoutMS=5000)
db = client["AlAhram_DB"]

users = db["users"]
stock = db["stock"]
cards = db["cards"]
transactions = db["transactions"]
counters = db["counters"]
admins_db = db["admins"]

SHEET_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbzPrw8oANq8Aek6O6URoTU0kDVjb1ZtoVdYkhpqAqM6Nuws4ZmcPRC9JtoNZvWoMzUb/exec"

bot = telebot.TeleBot(CUSTOMER_BOT_TOKEN)

# ==========================================
# الدوال المساعدة والفواتير
# ==========================================
def safe_str(text):
    if text is None: return "بدون"
    res = str(text)
    for char in ["_", "*", "`", "[", "]", "(", ")", "~", ">", "#", "+", "-", "=", "|", "{", "}", ".", "!"]:
        res = res.replace(char, " ")
    return res.strip()

def is_valid_val(val):
    if val is None: return False
    if str(val).strip().lower() in ["", "none", "null", "بدون", "nan"]: return False
    return True

def get_all_admins():
    admins = list(ADMIN_IDS)
    for a in admins_db.find():
        if a["_id"] not in admins: admins.append(a["_id"])
    return admins

def get_next_order_id():
    doc = counters.find_one_and_update({"_id": "order_id"}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER)
    return doc["seq"]

def check_user_access(uid):
    u = users.find_one({"_id": uid})
    if not u:
        bot.send_message(uid, "⚠️ النظام مغلق. أرسل /start للتسجيل أولاً.")
        return None
    if not u.get("name"):
        msg = bot.send_message(uid, "يرجى كتابة اسمك الكريم أولاً للاستمرار:")
        bot.register_next_step_handler(msg, process_name)
        return None
    if not u.get("phone"):
        bot.send_message(uid, "⚠️ يجب عليك مشاركة رقم هاتف حسابك أولاً.", reply_markup=contact_menu())
        return None
    if u.get("status") == "blocked":
        bot.send_message(uid, "تم حظرك من قبل الادارة وللاستفسار تواصل معنا")
        return None
    if u.get("status") != "active":
        bot.send_message(uid, "❌ حسابك قيد المراجعة. برجاء الانتظار حتى تفعيل الإدارة.")
        return None
    return u

def generate_customer_excel_file(items_data, order_id, dt_now, product_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoice_{order_id}"
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    info_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    
    has_serial = any(is_valid_val(d.get('serial')) for d in items_data)
    has_pin = any(is_valid_val(d.get('pin')) for d in items_data)
    has_opcode = any(is_valid_val(d.get('op_code')) for d in items_data)
    
    headers = ["م", "اسم المنتج", "كود الشحن"]
    if has_serial: headers.append("الرقم التسلسلي")
    if has_pin: headers.append("الرقم السري (PIN)")
    if has_opcode: headers.append("أوبريشن كود")
        
    last_col_idx = len(headers)
    last_col_letter = get_column_letter(last_col_idx)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = "شركة الأهرام للإتصالات والتقنية"
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=16); ws['A1'].fill = title_fill; ws['A1'].alignment = centered
    ws.row_dimensions[1].height = 35
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"فاتورة شراء رقم: #{order_id}   |   تاريخ العملية: {dt_now}"
    ws['A2'].fill = info_fill; ws['A2'].alignment = centered
    ws.row_dimensions[2].height = 25
    ws.append([]) 
    ws.append(headers)
    for cell in ws[4]:
        cell.font = white_font; cell.fill = header_fill; cell.alignment = centered
        
    for i, d in enumerate(items_data, 1):
        row_data = [i, product_name, str(d['code'])]
        if has_serial: row_data.append(str(d.get('serial', '')))
        if has_pin: row_data.append(str(d.get('pin', '')))
        if has_opcode: row_data.append(str(d.get('op_code', '')))
        ws.append(row_data)
        for cell in ws[ws.max_row]: cell.alignment = centered

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    stream.name = f"Invoice_{order_id}.xlsx"
    return stream

def menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🛒 شراء", "💳 شحن"); kb.add("👤 حسابي")
    return kb

def contact_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(types.KeyboardButton("إرسال رقم الهاتف 📱", request_contact=True))
    return kb

# ==========================================
# التسجيل وواجهة العميل
# ==========================================
@bot.message_handler(commands=['start'])
def handle_start(msg):
    uid = msg.chat.id
    u = users.find_one({"_id": uid})
    if not u:
        users.insert_one({"_id": uid, "name": None, "phone": None, "balance": 0.0, "status": "frozen", "tier": 1, "failed_attempts": 0, "join": datetime.datetime.now()})
        u = users.find_one({"_id": uid})

    if not u.get("name"):
        ask = bot.send_message(uid, "👋 مرحباً بك في شركة الأهرام للإتصالات.\n\nيرجى كتابة **اسمك بالكامل**:")
        bot.register_next_step_handler(ask, process_name)
    elif not u.get("phone"):
        bot.send_message(uid, f"أهلاً بك يا {safe_str(u.get('name'))}! يرجى تزويدنا برقم هاتفك.", reply_markup=contact_menu())
    else:
        if u.get("status") == "blocked": bot.send_message(uid, "تم حظرك من قبل الادارة وللاستفسار تواصل معنا")
        elif u.get("status") == "frozen": bot.send_message(uid, "حسابك قيد المراجعة. سيتم إشعارك فور التفعيل.")
        else: bot.send_message(uid, "أهلاً بك مجدداً في المتجر 🏪", reply_markup=menu())

def process_name(msg):
    if not msg.text or msg.text.startswith('/'):
        ask = bot.send_message(msg.chat.id, "يرجى كتابة اسم صحيح بدون رموز:")
        bot.register_next_step_handler(ask, process_name)
        return
    users.update_one({"_id": msg.chat.id}, {"$set": {"name": msg.text.strip()}})
    bot.send_message(msg.chat.id, f"تشرفنا بك!\nالآن يرجى مشاركة رقم هاتفك 📱", reply_markup=contact_menu())

@bot.message_handler(content_types=['contact'])
def handle_contact_sharing(msg):
    uid = msg.chat.id
    if msg.contact.user_id != uid: return bot.send_message(uid, "❌ يرجى مشاركة رقم هاتفك الخاص المرتبط بالحساب.", reply_markup=contact_menu())
    phone = msg.contact.phone_number
    if not phone.startswith('+'): phone = '+' + phone
    users.update_one({"_id": uid}, {"$set": {"phone": phone, "tier": 1}})
    u = users.find_one({"_id": uid})
    
    if u.get("status") == "frozen":
        bot.send_message(uid, "✅ تم استلام بياناتك بنجاح.\nحسابك قيد المراجعة.", reply_markup=types.ReplyKeyboardRemove())
        for admin_id in get_all_admins():
            try: bot.send_message(admin_id, f"🆕 **مستخدم جديد ينتظر التفعيل في بوت الإدارة!**\n📛 {safe_str(u.get('name'))}\n📱 `{phone}`")
            except: pass
    else: bot.send_message(uid, "✅ حسابك نشط وجاهز.", reply_markup=menu())

@bot.message_handler(func=lambda m: m.text == "👤 حسابي")
def show_account_info(msg):
    u = check_user_access(msg.chat.id)
    if not u: return
    text = f"👤 **بيانات حسابك**\n\n📛 الاسم: {safe_str(u.get('name'))}\n🆔 معرفك: `{msg.chat.id}`\n📱 الهاتف: `{u.get('phone')}`\n💰 رصيدك الحالي: **{round(float(u.get('balance', 0)), 2)}**\nحالة الحساب: {'نشط ✅' if u.get('status') == 'active' else 'قيد المراجعة ❄️'}"
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(types.InlineKeyboardButton("🛒 سجل المشتريات", callback_data="client_purchases"), types.InlineKeyboardButton("🧾 كشف حساب تفصيلي", callback_data="client_statement"))
    bot.send_message(msg.chat.id, text, parse_mode="Markdown", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data in ["client_purchases", "client_statement"])
def handle_client_reports(call):
    uid = call.message.chat.id
    u = check_user_access(uid)
    if not u: return bot.answer_callback_query(call.id, "حسابك غير مفعل.", show_alert=True)
    if call.data == "client_purchases":
        history = list(transactions.find({"uid": uid, "type": "شراء"}).sort("_id", -1).limit(10))
        if not history: return bot.answer_callback_query(call.id, "لا توجد مشتريات.", show_alert=True)
        txt = "🛒 **آخر 10 مشتريات:**\n\n"
        for t in history: txt += f"▪️ {t.get('date','')} | #{t.get('order_id','-')} | {t.get('item_name','')} (x{t.get('quantity',1)}) | {round(float(t.get('price',0)),2)}\n"
    elif call.data == "client_statement":
        history = list(transactions.find({"uid": uid}).sort("_id", -1).limit(20))
        if not history: return bot.answer_callback_query(call.id, "لا توجد حركات.", show_alert=True)
        txt = "🧾 **كشف حساب (آخر 20 حركة):**\n\n"
        for t in history:
            if t.get('type') == "شراء": txt += f"🔴 خصم | {t.get('date','')} | شراء {t.get('item_name','')} | -{round(float(t.get('price',0)),2)}\n"
            else: txt += f"🟢 إضافة | {t.get('date','')} | {t.get('type','')} | +{round(float(t.get('amount',0)),2)}\n"
    bot.send_message(uid, txt, parse_mode="Markdown")
    bot.answer_callback_query(call.id)

# ==========================================
# الشحن والتسوق
# ==========================================
@bot.message_handler(func=lambda m: m.text == "💳 شحن")
def ask_for_card_code(msg):
    if not check_user_access(msg.chat.id): return
    ask = bot.send_message(msg.chat.id, "يرجى إرسال كود شحن الرصيد:")
    bot.register_next_step_handler(ask, process_card_charging)

def process_card_charging(msg):
    if msg.text in ["🛒 شراء", "💳 شحن", "👤 حسابي"]: return bot.send_message(msg.chat.id, "تم إلغاء الشحن.")
    uid = msg.chat.id
    u = users.find_one({"_id": uid})
    if not u or u.get("status") != "active": return bot.send_message(uid, "❌ حسابك غير مفعل.")

    card = cards.find_one_and_update({"code": msg.text.strip(), "used": False}, {"$set": {"used": True, "used_by": uid, "used_at": datetime.datetime.now()}})
    if not card:
        attempts = u.get("failed_attempts", 0) + 1
        if attempts >= 5:
            users.update_one({"_id": uid}, {"$set": {"status": "frozen", "failed_attempts": 0}})
            bot.send_message(uid, "🚫 تم تجميد حسابك بسبب تكرار إدخال أكواد خاطئة.")
        else:
            users.update_one({"_id": uid}, {"$set": {"failed_attempts": attempts}})
            bot.send_message(uid, f"❌ الكود غير صحيح. متبقي لك {5 - attempts} محاولات.")
        return

    card_val = round(float(card["value"]), 2)
    new_bal = round(float(u.get("balance", 0.0)) + card_val, 2)
    users.update_one({"_id": uid}, {"$set": {"balance": new_bal, "failed_attempts": 0}})
    transactions.insert_one({"uid": uid, "user_name": u.get("name"), "type": "شحن كارت", "amount": card_val, "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
    bot.send_message(uid, f"✅ تم شحن حسابك بمبلغ {card_val}\n💰 رصيدك الجديد: {new_bal}")
    
    for admin_id in get_all_admins():
        try: bot.send_message(admin_id, f"💳 **شحن رصيد!**\n👤 العميل: {safe_str(u.get('name'))}\n💰 أضاف: {card_val}\n💵 أصبح رصيده: {new_bal}")
        except: pass

@bot.message_handler(func=lambda m: m.text == "🛒 شراء")
def start_shopping(msg):
    if not check_user_access(msg.chat.id): return
    cats = stock.distinct("category", {"sold": False})
    if not cats: return bot.send_message(msg.chat.id, "❌ لا توجد منتجات حالياً.")
    kb = types.InlineKeyboardMarkup()
    for c in cats: kb.add(types.InlineKeyboardButton(c, callback_data=f"cat_{c}"))
    bot.send_message(msg.chat.id, "اختر القسم:", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("cat_"))
def handle_category(call):
    cat_name = call.data.split("_", 1)[1]
    subs = stock.distinct("subcategory", {"category": cat_name, "sold": False})
    kb = types.InlineKeyboardMarkup()
    for s in subs: kb.add(types.InlineKeyboardButton(s, callback_data=f"sub_{cat_name}_{s}"))
    bot.edit_message_text(f"القسم: {cat_name}\nاختر الفئة:", call.message.chat.id, call.message.message_id, reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("sub_"))
def handle_subcategory(call):
    uid = call.message.chat.id
    u = users.find_one({"_id": uid})
    if not u or u.get("status") != "active": return
    
    tier = u.get("tier", 1)
    parts = call.data.split("_", 2)
    items = list(stock.find({"category": parts[1], "subcategory": parts[2], "sold": False}))
    if len(items) < 10: return bot.answer_callback_query(call.id, "⚠️ الكمية المتبقية أقل من 10.", show_alert=True)
        
    sample = items[0]
    p1 = float(sample.get("price_1", sample.get("price", 0)))
    final_price = round(p1 if tier == 1 else (float(sample.get("price_2", p1)) if tier == 2 else float(sample.get("price_3", p1))), 2)
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("🛒 تأكيد طلب شراء", callback_data=f"buy_{sample['_id']}"))
    bot.send_message(uid, f"📦 المنتج: **{sample['name']}**\n💰 السعر: **{final_price}**\n📊 المتوفر: {len(items)}\n\n⚠️ أقل كمية هي 10 ومضاعفاتها.", reply_markup=kb, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("buy_"))
def ask_qty(call):
    uid = call.message.chat.id
    u = users.find_one({"_id": uid})
    if not u or u.get("status") != "active": return
    
    item = stock.find_one({"_id": ObjectId(call.data.split("_")[1]), "sold": False})
    if not item: return bot.answer_callback_query(call.id, "❌ عذراً، المنتج مباع.")
    
    avail = stock.count_documents({"name": item['name'], "sold": False})
    ask = bot.send_message(uid, f"📦 {item['name']}\n📊 المتاح: {avail}\n\nيرجى إرسال الكمية (10، 20...):")
    bot.register_next_step_handler(ask, finalize_purchase, u, item, avail)

def finalize_purchase(msg, user_data, item_ref, avail_qty):
    if msg.text in ["🛒 شراء", "💳 شحن", "👤 حسابي"]: return bot.send_message(msg.chat.id, "تم إلغاء الشراء.")
    uid = msg.chat.id
    try:
        qty = int(msg.text.strip())
        if qty < 10 or qty % 10 != 0: return bot.send_message(uid, "❌ الكمية يجب أن تكون مضاعفات الـ 10.")
    except: return bot.send_message(uid, "❌ أرقام فقط.")
        
    if qty > avail_qty: return bot.send_message(uid, f"❌ المتاح فقط: {avail_qty}")

    tier = user_data.get("tier", 1)
    p1 = float(item_ref.get("price_1", item_ref.get("price", 0)))
    unit_price = p1 if tier == 1 else (float(item_ref.get("price_2", p1)) if tier == 2 else float(item_ref.get("price_3", p1)))
    total_cost = round(qty * unit_price, 2)
    user_balance = round(float(user_data.get("balance", 0.0)), 2)
    
    if user_balance < total_cost: return bot.send_message(uid, f"❌ رصيدك غير كافي.\nالمطلوب: {total_cost}\nرصيدك: {user_balance}")

    prod_name = item_ref['name']
    batch = list(stock.find({"name": prod_name, "sold": False}).limit(qty))
    if len(batch) < qty: return bot.send_message(uid, "❌ خطأ سحب، حاول مجدداً.")
        
    ids = [d['_id'] for d in batch]
    res = stock.update_many({"_id": {"$in": ids}, "sold": False}, {"$set": {"sold": True, "buyer_id": uid, "order_date": datetime.datetime.now()}})
    if res.modified_count != qty:
        stock.update_many({"_id": {"$in": ids}}, {"$set": {"sold": False}})
        return bot.send_message(uid, "❌ حدث تضارب.")

    users.update_one({"_id": uid}, {"$inc": {"balance": -total_cost}})
    order_id = get_next_order_id()
    dt_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    transactions.insert_one({"order_id": order_id, "uid": uid, "user_name": user_data.get("name"), "phone": user_data.get("phone"), "type": "شراء", "item_name": prod_name, "quantity": qty, "price": total_cost, "date": dt_str})
    if SHEET_WEBHOOK_URL:
        try: requests.post(SHEET_WEBHOOK_URL, json={"order_id": order_id, "date": dt_str, "phone": user_data.get('phone'), "item_name": f"{prod_name} (x{qty})", "price": total_cost}, timeout=3)
        except: pass
            
    bot.send_message(uid, f"✅ تم الشراء بنجاح!\n🧾 الفاتورة: #{order_id}\n💰 المخصوم: {total_cost}\n\nجاري إعداد الأكواد...")
    
    try:
        p_data = [{'code': d['code'], 'serial': d.get('serial',''), 'pin': d.get('pin',''), 'op_code': d.get('op_code','')} for d in batch]
        file_stream = generate_customer_excel_file(p_data, order_id, dt_str, prod_name)
        bot.send_document(uid, (f"Invoice_{order_id}.xlsx", file_stream.getvalue()), caption=f"📁 فاتورة الأكواد #{order_id}")
        
        admin_msg = f"🛒 **شراء جديد!**\nالفاتورة: #{order_id}\n👤 العميل: {safe_str(user_data.get('name'))}\n📦 {prod_name} (x{qty})\n💰 المبلغ: {total_cost}"
        for admin_id in get_all_admins():
            try:
                bot.send_message(admin_id, admin_msg)
                bot.send_document(admin_id, (f"Admin_Inv_{order_id}.xlsx", file_stream.getvalue()), caption="📁 نسخة للإدارة")
            except: pass
    except Exception as e:
        bot.send_message(uid, "⚠️ حدث خطأ أثناء إرسال الفاتورة لتليجرام، تم حفظ الأكواد لك، تواصل مع الإدارة.")

    rem = stock.count_documents({"name": prod_name, "sold": False})
    if rem <= 30:
        for admin_id in get_all_admins():
            try: bot.send_message(admin_id, f"⚠️ **نقص مخزون**\n`{prod_name}` المتبقي: {rem}")
            except: pass

# ==========================================
# تشغيل السيرفر
# ==========================================
app = Flask(__name__)
@app.route('/')
def health_check(): return "<h1>Customer Bot is Running 🚀</h1>"

def bot_polling():
    bot.remove_webhook()
    while True:
        try: bot.infinity_polling(timeout=30, long_polling_timeout=15)
        except: time.sleep(5)

if __name__ == "__main__":
    threading.Thread(target=bot_polling, daemon=True).start()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
